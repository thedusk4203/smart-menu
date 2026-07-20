from __future__ import annotations

import csv
import io

import pytest
from openpyxl import load_workbook

from app.modules.admin.normalization import (
    as_money as _as_money,
    as_optional_id as _as_optional_id,
    normalized_code as _normalized_code,
    normalized_tags as _normalized_tags,
)
from app.modules.admin.use_cases import AdminService


def test_import_code_is_normalized_and_rejects_unsafe_values():
    assert _normalized_code(" ing-001 ") == "ING-001"
    assert _normalized_code("") is None
    with pytest.raises(ValueError, match="code chỉ gồm"):
        _normalized_code("mã có khoảng trắng")


def test_optional_import_id_requires_positive_integer():
    assert _as_optional_id("12") == 12
    assert _as_optional_id("") is None
    with pytest.raises(ValueError, match="số nguyên dương"):
        _as_optional_id("0")


def test_import_tags_are_normalized_deduplicated_and_limited():
    assert _normalized_tags("  giàu đạm, nhanh , GIÀU ĐẠM ") == ["giàu đạm", "nhanh"]
    assert _normalized_tags(["rau củ", " rau   củ ", "ít dầu"]) == ["rau củ", "ít dầu"]
    with pytest.raises(ValueError, match="vượt quá 64"):
        _normalized_tags("x" * 65)


def test_import_money_accepts_raw_and_vietnamese_grouping():
    assert _as_money("1000000", "price", 2) == 1_000_000
    assert _as_money("1.000.000đ", "price", 2) == 1_000_000
    assert _as_money("1.234,5678", "price_per_default_unit", 2, max_fraction_digits=4) == 1234.5678
    assert _as_money("12.5", "price_per_default_unit", 2, max_fraction_digits=4) == 12.5
    with pytest.raises(ValueError, match="số nguyên đồng"):
        _as_money("12,5", "price", 2)
    with pytest.raises(ValueError, match="tối đa 4"):
        _as_money("1,12345", "price_per_default_unit", 2, max_fraction_digits=4)


def test_ingredient_csv_template_exposes_identity_columns():
    service = AdminService(None)  # type: ignore[arg-type]

    content, media_type, filename = service.import_template("ingredients", "csv")
    headers = next(csv.reader(io.StringIO(content.decode("utf-8-sig"))))

    assert media_type.startswith("text/csv")
    assert filename.endswith(".csv")
    assert headers[:3] == ["id", "code", "name"]
    assert headers[6] == "tags"
    assert "price_per_default_unit" in headers
    assert headers[-7:] == [
        "purchase_mode",
        "purchase_increment",
        "room_shelf_life_days",
        "fridge_shelf_life_days",
        "freezer_shelf_life_days",
        "shelf_life_source",
        "shelf_life_reviewed_at",
    ]


def test_dish_xlsx_template_includes_data_and_guide_sheets():
    service = AdminService(None)  # type: ignore[arg-type]

    content, media_type, filename = service.import_template("dishes", "xlsx")
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    headers = next(workbook["Dữ liệu"].iter_rows(values_only=True))

    assert media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert filename.endswith(".xlsx")
    assert workbook.sheetnames == ["Dữ liệu", "Hướng dẫn"]
    assert list(headers[:3]) == ["id", "code", "name"]
    assert "ingredients_json" in headers
    guide_rows = [row[0] for row in workbook["Hướng dẫn"].iter_rows(values_only=True)]
    assert any("Import nguyên liệu trước" in str(note) for note in guide_rows)


class _Row:
    def __init__(self, ingredient_id: int, name: str, default_unit: str) -> None:
        self.id = ingredient_id
        self.name = name
        self.default_unit = default_unit


class _Result:
    def __init__(self, row: _Row | None) -> None:
        self.row = row

    def first(self) -> _Row | None:
        return self.row


class _IngredientSession:
    def __init__(self) -> None:
        self.rows = [_Row(1, "Gạo", "g"), _Row(2, "Thịt gà", "g")]

    def execute(self, statement, params):  # type: ignore[no-untyped-def]
        if "id=:id" in str(statement):
            return _Result(next((row for row in self.rows if row.id == params["id"]), None))
        return _Result(next((row for row in self.rows if row.name.casefold() == params["name"]), None))


def test_dish_import_ingredients_require_existing_unique_positive_and_matching_units():
    service = AdminService(_IngredientSession())  # type: ignore[arg-type]
    parsed = service._parse_dish_ingredients(
        [{"ingredient_id": 1, "quantity": 150, "unit": "g"}], "Cơm gà"
    )
    assert parsed == [{"ingredient_id": 1, "quantity": 150.0, "unit": "g"}]

    with pytest.raises(ValueError, match="Không tìm thấy nguyên liệu"):
        service._parse_dish_ingredients([{"ingredient_id": 99, "quantity": 1, "unit": "g"}], "Cơm gà")
    with pytest.raises(ValueError, match="lớn hơn 0"):
        service._parse_dish_ingredients([{"ingredient_id": 1, "quantity": 0, "unit": "g"}], "Cơm gà")
    with pytest.raises(ValueError, match="không khớp đơn vị"):
        service._parse_dish_ingredients([{"ingredient_id": 1, "quantity": 1, "unit": "kg"}], "Cơm gà")
    with pytest.raises(ValueError, match="bị lặp"):
        service._parse_dish_ingredients([
            {"ingredient_id": 1, "quantity": 1, "unit": "g"},
            {"name": "Gạo", "quantity": 2, "unit": "g"},
        ], "Cơm gà")


def test_dish_import_accepts_bounded_leftover_flex_rules():
    service = AdminService(_IngredientSession())  # type: ignore[arg-type]

    parsed = service._parse_dish_ingredients(
        [{
            "ingredient_id": 1,
            "quantity": 150,
            "unit": "g",
            "max_extra_quantity": 20,
            "extra_step_quantity": 5,
        }],
        "Cơm gà",
    )

    assert parsed == [{
        "ingredient_id": 1,
        "quantity": 150.0,
        "unit": "g",
        "max_extra_quantity": 20.0,
        "extra_step_quantity": 5.0,
    }]
    with pytest.raises(ValueError, match="Bước tăng"):
        service._parse_dish_ingredients([{
            "ingredient_id": 1,
            "quantity": 150,
            "unit": "g",
            "max_extra_quantity": 5,
            "extra_step_quantity": 10,
        }], "Cơm gà")


def test_ingredient_export_csv_uses_import_headers_and_utf8_bom():
    service = AdminService(None)  # type: ignore[arg-type]
    content, media_type, filename = service._build_export_file("ingredients", "csv", [{
        "id": 7, "code": "ING-007", "name": "Đậu phụ", "food_group": "protein",
        "default_unit": "g", "grams_per_unit": 1, "tags": "giàu đạm, chay", "calories": 76,
        "protein_g": 8, "carbs_g": 2, "fat_g": 4, "fiber_g": 1,
        "price": 30000, "price_unit": "kg", "price_per_default_unit": 30,
        "source": "Siêu thị", "is_active": True,
    }])
    rows = list(csv.DictReader(io.StringIO(content.decode("utf-8-sig"))))

    assert media_type.startswith("text/csv")
    assert filename.startswith("smart-menu-ingredients-export-")
    assert rows == [{
        "id": "7", "code": "ING-007", "name": "Đậu phụ", "food_group": "protein",
        "default_unit": "g", "grams_per_unit": "1", "tags": "giàu đạm, chay", "calories": "76", "protein_g": "8",
        "carbs_g": "2", "fat_g": "4", "fiber_g": "1", "price": "30000",
        "price_unit": "kg", "price_per_default_unit": "30", "source": "Siêu thị", "is_active": "True",
        "purchase_mode": "", "purchase_increment": "", "room_shelf_life_days": "",
        "fridge_shelf_life_days": "", "freezer_shelf_life_days": "",
        "shelf_life_source": "", "shelf_life_reviewed_at": "",
    }]


def test_ingredient_export_round_trips_procurement_and_storage_metadata():
    service = AdminService(None)  # type: ignore[arg-type]
    content, _media_type, _filename = service._build_export_file("ingredients", "csv", [{
        "id": 8,
        "code": "ING-008",
        "name": "Ức gà",
        "food_group": "protein",
        "default_unit": "g",
        "purchase_mode": "regular",
        "purchase_increment": 100,
        "room_shelf_life_days": 0,
        "fridge_shelf_life_days": 2,
        "freezer_shelf_life_days": 30,
        "shelf_life_source": "Nhà cung cấp",
        "shelf_life_reviewed_at": "2026-07-01",
    }])

    row = next(csv.DictReader(io.StringIO(content.decode("utf-8-sig"))))
    assert row["purchase_mode"] == "regular"
    assert row["purchase_increment"] == "100"
    assert row["room_shelf_life_days"] == "0"
    assert row["fridge_shelf_life_days"] == "2"
    assert row["freezer_shelf_life_days"] == "30"
    assert row["shelf_life_source"] == "Nhà cung cấp"
    assert row["shelf_life_reviewed_at"] == "2026-07-01"


def test_dish_export_xlsx_is_import_compatible_and_has_guide():
    service = AdminService(None)  # type: ignore[arg-type]
    ingredients_json = '[{"ingredient_id": 7, "name": "Đậu phụ", "quantity": 120.0, "unit": "g"}]'
    content, media_type, filename = service._build_export_file("dishes", "xlsx", [{
        "id": 11, "code": "DISH-011", "name": "Đậu phụ sốt", "dish_type": "savory",
        "cooking_method": "braise", "description": None, "instructions": "Kho nhỏ lửa.",
        "tags": "đậu phụ, chay", "ingredients_json": ingredients_json, "is_active": True,
    }])
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    rows = list(workbook["Dữ liệu"].iter_rows(values_only=True))

    assert media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert filename.startswith("smart-menu-dishes-export-")
    assert workbook.sheetnames == ["Dữ liệu", "Hướng dẫn"]
    assert list(rows[0]) == ["id", "code", "name", "dish_type", "cooking_method", "description", "instructions", "tags", "ingredients_json", "is_active"]
    assert rows[1][8] == ingredients_json
    assert any("Import nguyên liệu trước" in str(row[0]) for row in workbook["Hướng dẫn"].iter_rows(values_only=True))
