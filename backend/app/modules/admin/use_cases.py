from __future__ import annotations

import csv
import io
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from sqlalchemy import text
from sqlmodel import Session

from app.core.exceptions import ConflictError, NotFoundError, ValidationAppError
from app.core.security import hash_password
from app.modules.admin.schemas import (
    AdminDishWrite,
    AdminIngredientWrite,
    AdminUserCreate,
)
from app.shared.enums import DishType, FoodGroup, UserRole


PRIVILEGED_ROLES = {UserRole.ADMIN.value, UserRole.SUPER_ADMIN.value}

# 1 g/ml là hợp lệ với nước và nhiều chất lỏng gần nước, nhưng thường chỉ là
# giá trị mặc định chưa được kiểm tra đối với dầu và sữa. Các đơn vị rời như
# quả/hộp/muỗng vẫn luôn cần quy đổi nếu hệ số còn bằng 1.
MISSING_CONVERSION_SQL = """i.grams_per_unit = 1
    AND LOWER(BTRIM(i.default_unit)) NOT IN ('g', 'gram', 'grams')
    AND (
        LOWER(BTRIM(i.default_unit)) NOT IN ('ml', 'milliliter', 'milliliters')
        OR i.food_group::text IN ('fat', 'dairy')
    )"""


def _row_dict(row) -> dict[str, Any]:
    return dict(row._mapping) if row is not None else {}


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=str)


def _normalized(value: str) -> str:
    return " ".join(value.casefold().split())


def _normalized_tags(value: Any) -> list[str]:
    raw_tags = value if isinstance(value, list) else str(value or "").split(",")
    tags: list[str] = []
    seen: set[str] = set()
    for raw_tag in raw_tags:
        tag = " ".join(str(raw_tag).split())
        if not tag:
            continue
        if len(tag) > 64:
            raise ValueError(f"Thẻ '{tag}' vượt quá 64 ký tự")
        key = tag.casefold()
        if key not in seen:
            seen.add(key)
            tags.append(tag)
    return tags


CODE_PATTERN = re.compile(r"^[A-Z0-9][A-Z0-9._-]{0,63}$")


def _normalized_code(value: Any) -> str | None:
    if value is None or str(value).strip() == "":
        return None
    code = str(value).strip().upper()
    if not CODE_PATTERN.fullmatch(code):
        raise ValueError("code chỉ gồm chữ in hoa, số, dấu chấm, gạch dưới hoặc gạch ngang")
    return code


def _as_optional_id(value: Any) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        result = int(str(value).strip())
    except (TypeError, ValueError) as exc:
        raise ValueError("id phải là số nguyên dương") from exc
    if result < 1:
        raise ValueError("id phải là số nguyên dương")
    return result


def _enum_value(value: Any) -> str:
    return value.value if hasattr(value, "value") else str(value)


def _as_bool(value: Any, default: bool = True) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().casefold() in {"1", "true", "yes", "y", "có", "co"}


def _as_float(value: Any, field: str, row_number: int, *, default: float | None = None) -> float | None:
    if value is None or str(value).strip() == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field} phải là số") from exc


def _as_money(
    value: Any,
    field: str,
    row_number: int,
    *,
    default: float | None = None,
    max_fraction_digits: int = 0,
) -> float | None:
    """Đọc số tiền thô hoặc định dạng Việt Nam từ CSV/XLSX thành số chuẩn."""
    if value is None or str(value).strip() == "":
        return default
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        result = float(value)
    else:
        raw = str(value).strip().replace("đ", "").replace("Đ", "")
        raw = re.sub(r"\s+", "", raw)
        if not raw or not re.fullmatch(r"[0-9.,]+", raw):
            raise ValueError(f"{field} phải là số tiền hợp lệ")
        fraction = ""
        if "," in raw:
            if raw.count(",") != 1:
                raise ValueError(f"{field} phải là số tiền hợp lệ")
            integer, fraction = raw.split(",")
            integer = integer.replace(".", "")
        elif re.fullmatch(r"\d{1,3}(?:\.\d{3})+", raw):
            integer = raw.replace(".", "")
        elif re.fullmatch(r"\d+", raw):
            integer = raw
        elif re.fullmatch(r"\d+\.\d+", raw):
            integer, fraction = raw.split(".")
        else:
            raise ValueError(f"{field} phải là số tiền hợp lệ")
        if not integer.isdigit() or (fraction and not fraction.isdigit()):
            raise ValueError(f"{field} phải là số tiền hợp lệ")
        if fraction and max_fraction_digits == 0:
            raise ValueError(f"{field} phải là số nguyên đồng")
        if len(fraction) > max_fraction_digits:
            raise ValueError(f"{field} chỉ được có tối đa {max_fraction_digits} số lẻ")
        result = float(f"{integer}.{fraction}" if fraction else integer)
    if result < 0:
        raise ValueError(f"{field} không được âm")
    if max_fraction_digits == 0 and not result.is_integer():
        raise ValueError(f"{field} phải là số nguyên đồng")
    return result


class AdminService:
    """Ứng dụng quản trị tập trung cho dashboard, dữ liệu canonical và import.

    Các phép ghi dùng cùng một Session để công thức, giá và audit log được commit
    nguyên tử. Public read APIs vẫn nằm trong module domain hiện có.
    """

    def __init__(self, session: Session) -> None:
        self.session = session

    def _validate_catalog_tags(self, tags: list[str], entity_type: str = "dish") -> None:
        names = _normalized_tags(tags)
        if not names:
            return
        rows = self.session.execute(
            text("""SELECT name FROM tag_catalog
                    WHERE entity_type=:entity_type AND is_active=TRUE AND name = ANY(:names)"""),
            {"entity_type": entity_type, "names": names},
        ).fetchall()
        found = {row.name for row in rows}
        missing = [name for name in names if name not in found]
        if missing:
            raise ValidationAppError("Thẻ chưa có trong danh mục hoặc đã ngừng dùng: " + ", ".join(missing))

    def _ensure_catalog_tags(self, tags: list[str], entity_type: str) -> None:
        for name in _normalized_tags(tags):
            existing = self.session.execute(
                text("""SELECT id FROM tag_catalog
                        WHERE entity_type=:entity_type AND LOWER(name)=LOWER(:name)
                        FOR UPDATE"""),
                {"entity_type": entity_type, "name": name},
            ).first()
            if existing:
                self.session.execute(
                    text("UPDATE tag_catalog SET is_active=TRUE, updated_at=NOW() WHERE id=:id"),
                    {"id": existing.id},
                )
            else:
                self.session.execute(
                    text("INSERT INTO tag_catalog (entity_type, name) VALUES (:entity_type, :name)"),
                    {"entity_type": entity_type, "name": name},
                )

    def _audit(
        self,
        actor_id: int,
        action: str,
        entity_type: str,
        entity_id: int | None,
        before: Any = None,
        after: Any = None,
    ) -> None:
        self.session.execute(
            text(
                """INSERT INTO audit_logs
                   (actor_user_id, action, entity_type, entity_id, before_data, after_data)
                   VALUES (:actor, :action, :etype, :eid,
                           CAST(:before AS jsonb), CAST(:after AS jsonb))"""
            ),
            {
                "actor": actor_id,
                "action": action,
                "etype": entity_type,
                "eid": entity_id,
                "before": _json(before) if before is not None else None,
                "after": _json(after) if after is not None else None,
            },
        )

    # ------------------------------------------------------------------ dashboard
    def dashboard(self) -> dict[str, Any]:
        row = self.session.execute(
            text(
                f"""SELECT
                    (SELECT COUNT(*) FROM users) AS users_total,
                    (SELECT COUNT(*) FROM users WHERE is_active) AS users_active,
                    (SELECT COUNT(*) FROM users WHERE NOT is_active) AS users_locked,
                    (SELECT COUNT(*) FROM ingredients) AS ingredients_total,
                    (SELECT COUNT(*) FROM ingredients WHERE is_active) AS ingredients_active,
                    (SELECT COUNT(*) FROM dishes) AS dishes_total,
                    (SELECT COUNT(*) FROM v_dish_candidates) AS planner_ready_dishes,
                    (SELECT COUNT(*) FROM v_dish_candidates WHERE dish_type = 'breakfast') AS breakfast_count,
                    (SELECT COUNT(*) FROM v_dish_candidates WHERE dish_type = 'staple') AS staple_count,
                    (SELECT COUNT(*) FROM v_dish_candidates WHERE dish_type = 'savory') AS savory_count,
                    (SELECT COUNT(*) FROM v_dish_candidates WHERE dish_type = 'vegetable_side') AS vegetable_count,
                    (SELECT COUNT(*) FROM v_dish_candidates WHERE dish_type = 'soup') AS soup_count,
                    (SELECT COUNT(*) FROM ingredients i WHERE NOT EXISTS (
                        SELECT 1 FROM price_snapshots p
                        WHERE p.ingredient_id = i.id AND p.price_per_default_unit IS NOT NULL
                    )) AS missing_price,
                    (SELECT COUNT(*) FROM ingredients i WHERE NOT EXISTS (
                        SELECT 1 FROM nutrition_facts n WHERE n.ingredient_id = i.id
                    )) AS missing_nutrition,
                    (SELECT COUNT(*) FROM ingredients i
                     WHERE {MISSING_CONVERSION_SQL}) AS missing_conversion,
                    (SELECT COUNT(*) FROM dishes d WHERE
                        NOT EXISTS (SELECT 1 FROM dish_ingredients di WHERE di.dish_id = d.id)
                        OR EXISTS (
                            SELECT 1 FROM dish_ingredients di
                            LEFT JOIN nutrition_facts n ON n.ingredient_id = di.ingredient_id
                            WHERE di.dish_id = d.id AND n.id IS NULL
                        )
                        OR EXISTS (
                            SELECT 1 FROM dish_ingredients di
                            WHERE di.dish_id = d.id AND NOT EXISTS (
                                SELECT 1 FROM price_snapshots p
                                WHERE p.ingredient_id = di.ingredient_id
                                  AND p.price_per_default_unit IS NOT NULL
                            )
                        )
                    ) AS incomplete_dishes,
                    (SELECT COALESCE(SUM(group_size - 1), 0) FROM (
                        SELECT COUNT(*) AS group_size FROM (
                            SELECT 'ingredient'::text AS entity_type,
                                   LOWER(REGEXP_REPLACE(BTRIM(name), '\\s+', ' ', 'g')) AS n
                            FROM ingredients
                            UNION ALL
                            SELECT 'dish'::text,
                                   LOWER(REGEXP_REPLACE(BTRIM(name), '\\s+', ' ', 'g'))
                            FROM dishes
                        ) names GROUP BY entity_type, n HAVING COUNT(*) > 1
                    ) duplicate_groups) AS duplicate_names,
                    (SELECT MAX(created_at) FROM import_jobs) AS last_import_at"""
            )
        ).first()
        data = _row_dict(row)
        data["open_quality_issues"] = (
            data["missing_price"]
            + data["missing_nutrition"]
            + data["missing_conversion"]
            + data["incomplete_dishes"]
            + data["duplicate_names"]
        )
        return data

    # ------------------------------------------------------------------ users
    def list_users(
        self,
        search: str | None,
        role: str | None,
        is_active: bool | None,
        limit: int,
        offset: int,
    ) -> dict[str, Any]:
        where = ["1=1"]
        params: dict[str, Any] = {"limit": limit, "offset": offset}
        if search:
            where.append("(u.email ILIKE :search OR p.full_name ILIKE :search)")
            params["search"] = f"%{search.strip()}%"
        if role:
            where.append("u.role::text = :role")
            params["role"] = role
        if is_active is not None:
            where.append("u.is_active = :active")
            params["active"] = is_active
        clause = " AND ".join(where)
        total = self.session.execute(
            text(
                f"""SELECT COUNT(*) FROM users u
                    LEFT JOIN user_profiles p ON p.user_id = u.id WHERE {clause}"""
            ),
            params,
        ).scalar_one()
        rows = self.session.execute(
            text(
                f"""SELECT u.id, u.email, p.full_name, u.role, u.is_active,
                           u.created_at, u.updated_at
                    FROM users u LEFT JOIN user_profiles p ON p.user_id = u.id
                    WHERE {clause}
                    ORDER BY u.created_at DESC, u.id DESC
                    LIMIT :limit OFFSET :offset"""
            ),
            params,
        ).fetchall()
        return {"items": [_row_dict(r) for r in rows], "total": total, "limit": limit, "offset": offset}

    def get_user(self, user_id: int) -> dict[str, Any]:
        row = self.session.execute(
            text(
                """SELECT u.id, u.email, p.full_name, u.role, u.is_active,
                          u.created_at, u.updated_at
                   FROM users u LEFT JOIN user_profiles p ON p.user_id = u.id
                   WHERE u.id = :id"""
            ),
            {"id": user_id},
        ).first()
        if not row:
            raise NotFoundError("Không tìm thấy người dùng")
        return _row_dict(row)

    def create_user(self, data: AdminUserCreate, actor_id: int) -> dict[str, Any]:
        if self.session.execute(text("SELECT 1 FROM users WHERE LOWER(email) = LOWER(:email)"), {"email": data.email}).first():
            raise ConflictError("Email đã được sử dụng")
        row = self.session.execute(
            text(
                """INSERT INTO users (email, hashed_password, role)
                   VALUES (:email, :password, CAST(:role AS user_role)) RETURNING id"""
            ),
            {"email": data.email, "password": hash_password(data.password), "role": data.role.value},
        ).first()
        user_id = row.id
        self.session.execute(
            text("INSERT INTO user_profiles (user_id, full_name) VALUES (:id, :name)"),
            {"id": user_id, "name": data.full_name},
        )
        after = self.get_user(user_id)
        self._audit(actor_id, "create", "user", user_id, after=after)
        self.session.commit()
        return after

    def update_user_role(self, user_id: int, role: UserRole, actor_id: int) -> dict[str, Any]:
        before = self.get_user(user_id)
        if user_id == actor_id and role.value != _enum_value(before["role"]):
            raise ValidationAppError("Bạn không thể tự thay đổi vai trò của chính mình")
        current_role = _enum_value(before["role"])
        if current_role in PRIVILEGED_ROLES and role.value not in PRIVILEGED_ROLES:
            count = self.session.execute(
                text("SELECT COUNT(*) FROM users WHERE role::text IN ('admin', 'super_admin') AND is_active")
            ).scalar_one()
            if count <= 1:
                raise ValidationAppError("Phải giữ lại ít nhất một quản trị viên hệ thống đang hoạt động")
        self.session.execute(
            text("UPDATE users SET role = CAST(:role AS user_role) WHERE id = :id"),
            {"role": role.value, "id": user_id},
        )
        after = self.get_user(user_id)
        self._audit(actor_id, "change_role", "user", user_id, before, after)
        self.session.commit()
        return after

    def update_user_status(self, user_id: int, is_active: bool, actor_id: int) -> dict[str, Any]:
        before = self.get_user(user_id)
        if user_id == actor_id and not is_active:
            raise ValidationAppError("Bạn không thể tự khóa tài khoản của chính mình")
        if _enum_value(before["role"]) in PRIVILEGED_ROLES and not is_active:
            count = self.session.execute(
                text("SELECT COUNT(*) FROM users WHERE role::text IN ('admin', 'super_admin') AND is_active")
            ).scalar_one()
            if count <= 1:
                raise ValidationAppError("Không thể khóa quản trị viên hệ thống cuối cùng")
        self.session.execute(
            text("UPDATE users SET is_active = :active WHERE id = :id"),
            {"active": is_active, "id": user_id},
        )
        after = self.get_user(user_id)
        self._audit(actor_id, "unlock" if is_active else "lock", "user", user_id, before, after)
        self.session.commit()
        return after

    # ------------------------------------------------------------------ ingredients
    def _ingredient_from_db(self, ingredient_id: int) -> dict[str, Any]:
        row = self.session.execute(
            text(
                f"""SELECT i.id, i.name, i.food_group, i.default_unit, i.grams_per_unit,
                          i.tags, i.is_active, n.calories, n.protein_g, n.carbs_g, n.fat_g,
                          n.fiber_g, p.price AS latest_price, p.unit AS price_unit,
                          p.price_per_default_unit AS latest_price_per_unit,
                          p.source AS price_source, p.recorded_at AS price_recorded_at,
                          i.created_at, i.updated_at,
                          (p.id IS NULL OR p.price_per_default_unit IS NULL) AS missing_price,
                          (n.id IS NULL) AS missing_nutrition,
                          ({MISSING_CONVERSION_SQL}) AS missing_conversion
                   FROM ingredients i
                   LEFT JOIN nutrition_facts n ON n.ingredient_id = i.id
                   LEFT JOIN LATERAL (
                       SELECT id, price, unit, price_per_default_unit, source, recorded_at
                       FROM price_snapshots WHERE ingredient_id = i.id
                       ORDER BY recorded_at DESC LIMIT 1
                   ) p ON TRUE
                   WHERE i.id = :id"""
            ),
            {"id": ingredient_id},
        ).first()
        if not row:
            raise NotFoundError("Không tìm thấy nguyên liệu")
        data = _row_dict(row)
        for key in (
            "grams_per_unit", "calories", "protein_g", "carbs_g", "fat_g",
            "fiber_g", "latest_price", "latest_price_per_unit",
        ):
            if data.get(key) is not None:
                data[key] = float(data[key])
        return data

    def list_ingredients(
        self,
        search: str | None,
        food_group: str | None,
        status: str | None,
        quality: str | None,
        limit: int,
        offset: int,
    ) -> dict[str, Any]:
        where = ["1=1"]
        params: dict[str, Any] = {"limit": limit, "offset": offset}
        if search:
            where.append("i.name ILIKE :search")
            params["search"] = f"%{search.strip()}%"
        if food_group:
            where.append("i.food_group::text = :food_group")
            params["food_group"] = food_group
        if status == "active":
            where.append("i.is_active")
        elif status == "inactive":
            where.append("NOT i.is_active")
        if quality == "missing_price":
            where.append("(p.id IS NULL OR p.price_per_default_unit IS NULL)")
        elif quality == "missing_nutrition":
            where.append("n.id IS NULL")
        elif quality == "missing_conversion":
            where.append(MISSING_CONVERSION_SQL)
        clause = " AND ".join(where)
        joins = """LEFT JOIN nutrition_facts n ON n.ingredient_id = i.id
                   LEFT JOIN LATERAL (
                       SELECT id, price, unit, price_per_default_unit, source, recorded_at
                       FROM price_snapshots WHERE ingredient_id = i.id
                       ORDER BY recorded_at DESC LIMIT 1
                   ) p ON TRUE"""
        total = self.session.execute(
            text(f"SELECT COUNT(*) FROM ingredients i {joins} WHERE {clause}"), params
        ).scalar_one()
        rows = self.session.execute(
            text(
                f"""SELECT i.id, i.name, i.food_group, i.default_unit, i.grams_per_unit,
                           i.tags, i.is_active, n.calories, n.protein_g, n.carbs_g, n.fat_g,
                           n.fiber_g, p.price AS latest_price, p.unit AS price_unit,
                           p.price_per_default_unit AS latest_price_per_unit,
                           p.source AS price_source, p.recorded_at AS price_recorded_at,
                           i.created_at, i.updated_at,
                           (p.id IS NULL OR p.price_per_default_unit IS NULL) AS missing_price,
                           (n.id IS NULL) AS missing_nutrition,
                           ({MISSING_CONVERSION_SQL}) AS missing_conversion
                    FROM ingredients i {joins} WHERE {clause}
                    ORDER BY i.updated_at DESC, i.name
                    LIMIT :limit OFFSET :offset"""
            ),
            params,
        ).fetchall()
        items = []
        for row in rows:
            data = _row_dict(row)
            for key in (
                "grams_per_unit", "calories", "protein_g", "carbs_g", "fat_g",
                "fiber_g", "latest_price", "latest_price_per_unit",
            ):
                if data.get(key) is not None:
                    data[key] = float(data[key])
            items.append(data)
        return {"items": items, "total": total, "limit": limit, "offset": offset}

    def export_ingredients(
        self,
        output_format: str,
        search: str | None,
        food_group: str | None,
        status: str | None,
        quality: str | None,
    ) -> tuple[bytes, str, str]:
        where = ["1=1"]
        params: dict[str, Any] = {}
        if search:
            where.append("i.name ILIKE :search")
            params["search"] = f"%{search.strip()}%"
        if food_group:
            where.append("i.food_group::text = :food_group")
            params["food_group"] = food_group
        if status == "active":
            where.append("i.is_active")
        elif status == "inactive":
            where.append("NOT i.is_active")
        if quality == "missing_price":
            where.append("(p.id IS NULL OR p.price_per_default_unit IS NULL)")
        elif quality == "missing_nutrition":
            where.append("n.id IS NULL")
        elif quality == "missing_conversion":
            where.append(MISSING_CONVERSION_SQL)
        joins = """LEFT JOIN nutrition_facts n ON n.ingredient_id = i.id
                   LEFT JOIN LATERAL (
                       SELECT id, price, unit, price_per_default_unit, source, recorded_at
                       FROM price_snapshots WHERE ingredient_id = i.id
                       ORDER BY recorded_at DESC LIMIT 1
                   ) p ON TRUE"""
        rows = self.session.execute(
            text(
                f"""SELECT i.id, i.code, i.name, i.food_group, i.default_unit, i.grams_per_unit, i.tags,
                           n.calories, n.protein_g, n.carbs_g, n.fat_g, n.fiber_g,
                           p.price, p.unit AS price_unit, p.price_per_default_unit, p.source, i.is_active
                    FROM ingredients i {joins}
                    WHERE {' AND '.join(where)}
                    ORDER BY i.updated_at DESC, i.name"""
            ),
            params,
        ).fetchall()
        records = []
        for row in rows:
            data = _row_dict(row)
            records.append({
                "id": data["id"], "code": data["code"], "name": data["name"],
                "food_group": _enum_value(data["food_group"]), "default_unit": data["default_unit"],
                "grams_per_unit": float(data["grams_per_unit"]),
                "tags": ", ".join(str(tag).strip() for tag in (data.get("tags") or []) if str(tag).strip()),
                "calories": float(data["calories"]) if data["calories"] is not None else None,
                "protein_g": float(data["protein_g"]) if data["protein_g"] is not None else None,
                "carbs_g": float(data["carbs_g"]) if data["carbs_g"] is not None else None,
                "fat_g": float(data["fat_g"]) if data["fat_g"] is not None else None,
                "fiber_g": float(data["fiber_g"]) if data["fiber_g"] is not None else None,
                "price": float(data["price"]) if data["price"] is not None else None,
                "price_unit": data["price_unit"],
                "price_per_default_unit": float(data["price_per_default_unit"])
                if data["price_per_default_unit"] is not None else None,
                "source": data["source"], "is_active": data["is_active"],
            })
        return self._build_export_file("ingredients", output_format, records)

    def save_ingredient(
        self,
        data: AdminIngredientWrite,
        actor_id: int,
        ingredient_id: int | None = None,
    ) -> dict[str, Any]:
        duplicate = self.session.execute(
            text(
                """SELECT id FROM ingredients
                   WHERE LOWER(REGEXP_REPLACE(BTRIM(name), '\\s+', ' ', 'g')) = :name
                     AND (:id IS NULL OR id <> :id)"""
            ),
            {"name": _normalized(data.name), "id": ingredient_id},
        ).first()
        if duplicate:
            raise ConflictError("Tên nguyên liệu đã tồn tại")
        before = self._ingredient_from_db(ingredient_id) if ingredient_id else None
        if ingredient_id is None:
            row = self.session.execute(
                text(
                    """INSERT INTO ingredients
                       (name, food_group, default_unit, grams_per_unit, is_active)
                       VALUES (:name, CAST(:group AS food_group), :unit, :grams, :active)
                       RETURNING id"""
                ),
                {
                    "name": data.name, "group": data.food_group.value,
                    "unit": data.default_unit, "grams": data.grams_per_unit,
                    "active": data.is_active,
                },
            ).first()
            ingredient_id = row.id
            action = "create"
        else:
            self.session.execute(
                text(
                    """UPDATE ingredients SET name = :name,
                           food_group = CAST(:group AS food_group), default_unit = :unit,
                           grams_per_unit = :grams, is_active = :active
                       WHERE id = :id"""
                ),
                {
                    "id": ingredient_id, "name": data.name, "group": data.food_group.value,
                    "unit": data.default_unit, "grams": data.grams_per_unit,
                    "active": data.is_active,
                },
            )
            action = "update"
        if data.nutrition is not None:
            nutrition = data.nutrition.model_dump()
            self.session.execute(
                text(
                    """INSERT INTO nutrition_facts
                       (ingredient_id, calories, protein_g, carbs_g, fat_g, fiber_g)
                       VALUES (:ingredient_id, :calories, :protein_g, :carbs_g, :fat_g, :fiber_g)
                       ON CONFLICT (ingredient_id) DO UPDATE SET
                           calories = EXCLUDED.calories, protein_g = EXCLUDED.protein_g,
                           carbs_g = EXCLUDED.carbs_g, fat_g = EXCLUDED.fat_g,
                           fiber_g = EXCLUDED.fiber_g"""
                ),
                {"ingredient_id": ingredient_id, **nutrition},
            )
        if data.price is not None:
            price = data.price.model_dump()
            self.session.execute(
                text(
                    """INSERT INTO price_snapshots
                       (ingredient_id, price, unit, price_per_default_unit, source, recorded_at)
                       VALUES (:ingredient_id, :price, :unit, :price_per_default_unit,
                               :source, COALESCE(:recorded_at, NOW()))"""
                ),
                {"ingredient_id": ingredient_id, **price},
            )
        after = self._ingredient_from_db(ingredient_id)
        self._audit(actor_id, action, "ingredient", ingredient_id, before, after)
        self.session.commit()
        return after

    def set_ingredient_active(self, ingredient_id: int, active: bool, actor_id: int) -> dict[str, Any]:
        before = self._ingredient_from_db(ingredient_id)
        self.session.execute(
            text("UPDATE ingredients SET is_active = :active WHERE id = :id"),
            {"active": active, "id": ingredient_id},
        )
        after = self._ingredient_from_db(ingredient_id)
        self._audit(actor_id, "restore" if active else "deactivate", "ingredient", ingredient_id, before, after)
        self.session.commit()
        return after

    def delete_ingredient(self, ingredient_id: int, actor_id: int) -> None:
        before = self._ingredient_from_db(ingredient_id)
        references = self.session.execute(
            text(
                """SELECT 'công thức món' AS label FROM dish_ingredients WHERE ingredient_id = :id LIMIT 1
                   UNION ALL
                   SELECT 'món ăn' AS label FROM meal_ingredients WHERE ingredient_id = :id LIMIT 1
                   UNION ALL
                   SELECT 'danh sách đi chợ' AS label FROM shopping_lists WHERE ingredient_id = :id LIMIT 1"""
            ),
            {"id": ingredient_id},
        ).fetchall()
        if references:
            labels = ", ".join(row.label for row in references)
            raise ConflictError(f"Không thể xóa nguyên liệu vì đang được dùng trong: {labels}. Hãy gỡ liên kết hoặc ẩn nguyên liệu.")
        self.session.execute(text("DELETE FROM ingredients WHERE id = :id"), {"id": ingredient_id})
        self._audit(actor_id, "delete", "ingredient", ingredient_id, before, None)
        self.session.commit()

    # ------------------------------------------------------------------ dishes
    def _dish_base_query(self) -> str:
        return """SELECT d.id, d.name, d.dish_type, d.cooking_method, d.description,
                         d.instructions, d.tags, d.is_active,
                         COALESCE(v.total_calories, 0) AS total_calories,
                         COALESCE(v.total_protein_g, 0) AS total_protein_g,
                         COALESCE(v.total_carbs_g, 0) AS total_carbs_g,
                         COALESCE(v.total_fat_g, 0) AS total_fat_g,
                         COALESCE(v.estimated_cost, 0) AS estimated_cost,
                         COALESCE(v.ingredient_count, 0) AS ingredient_count,
                         (COALESCE(v.ingredient_count, 0) = 0) AS missing_recipe,
                         EXISTS (
                           SELECT 1 FROM dish_ingredients di
                           WHERE di.dish_id = d.id AND NOT EXISTS (
                             SELECT 1 FROM price_snapshots p
                             WHERE p.ingredient_id = di.ingredient_id
                               AND p.price_per_default_unit IS NOT NULL
                           )
                         ) AS missing_price,
                         EXISTS (
                           SELECT 1 FROM dish_ingredients di
                           LEFT JOIN nutrition_facts n ON n.ingredient_id = di.ingredient_id
                           WHERE di.dish_id = d.id AND n.id IS NULL
                         ) AS missing_nutrition,
                         d.created_at, d.updated_at
                  FROM dishes d LEFT JOIN v_dishes_full v ON v.id = d.id"""

    def _dish_from_db(self, dish_id: int, include_ingredients: bool = True) -> dict[str, Any]:
        row = self.session.execute(
            text(self._dish_base_query() + " WHERE d.id = :id"), {"id": dish_id}
        ).first()
        if not row:
            raise NotFoundError("Không tìm thấy món thành phần")
        data = _row_dict(row)
        for key in (
            "total_calories", "total_protein_g", "total_carbs_g",
            "total_fat_g", "estimated_cost",
        ):
            data[key] = float(data[key] or 0)
        data["ingredients"] = []
        if include_ingredients:
            rows = self.session.execute(
                text(
                    """SELECT di.ingredient_id, i.name, di.quantity, di.unit,
                              NOT EXISTS (SELECT 1 FROM price_snapshots p
                                  WHERE p.ingredient_id = i.id
                                    AND p.price_per_default_unit IS NOT NULL) AS missing_price,
                              NOT EXISTS (SELECT 1 FROM nutrition_facts n
                                  WHERE n.ingredient_id = i.id) AS missing_nutrition
                       FROM dish_ingredients di JOIN ingredients i ON i.id = di.ingredient_id
                       WHERE di.dish_id = :id ORDER BY i.name"""
                ),
                {"id": dish_id},
            ).fetchall()
            data["ingredients"] = [
                {**_row_dict(r), "quantity": float(r.quantity)} for r in rows
            ]
        return data

    def list_dishes(
        self,
        search: str | None,
        dish_type: str | None,
        status: str | None,
        quality: str | None,
        limit: int,
        offset: int,
    ) -> dict[str, Any]:
        where = ["1=1"]
        params: dict[str, Any] = {"limit": limit, "offset": offset}
        if search:
            where.append("d.name ILIKE :search")
            params["search"] = f"%{search.strip()}%"
        if dish_type:
            where.append("d.dish_type::text = :dtype")
            params["dtype"] = dish_type
        if status == "active":
            where.append("d.is_active")
        elif status == "inactive":
            where.append("NOT d.is_active")
        if quality == "missing_recipe":
            where.append("COALESCE(v.ingredient_count, 0) = 0")
        elif quality == "missing_price":
            where.append("EXISTS (SELECT 1 FROM dish_ingredients di WHERE di.dish_id = d.id AND NOT EXISTS (SELECT 1 FROM price_snapshots p WHERE p.ingredient_id = di.ingredient_id AND p.price_per_default_unit IS NOT NULL))")
        elif quality == "missing_nutrition":
            where.append("EXISTS (SELECT 1 FROM dish_ingredients di LEFT JOIN nutrition_facts n ON n.ingredient_id = di.ingredient_id WHERE di.dish_id = d.id AND n.id IS NULL)")
        clause = " AND ".join(where)
        total = self.session.execute(
            text(f"SELECT COUNT(*) FROM dishes d LEFT JOIN v_dishes_full v ON v.id = d.id WHERE {clause}"),
            params,
        ).scalar_one()
        rows = self.session.execute(
            text(self._dish_base_query() + f" WHERE {clause} ORDER BY d.updated_at DESC, d.name LIMIT :limit OFFSET :offset"),
            params,
        ).fetchall()
        items = []
        for row in rows:
            data = _row_dict(row)
            for key in ("total_calories", "total_protein_g", "total_carbs_g", "total_fat_g", "estimated_cost"):
                data[key] = float(data[key] or 0)
            data["ingredients"] = []
            items.append(data)
        return {"items": items, "total": total, "limit": limit, "offset": offset}

    def export_dishes(
        self,
        output_format: str,
        search: str | None,
        dish_type: str | None,
        status: str | None,
        quality: str | None,
    ) -> tuple[bytes, str, str]:
        where = ["1=1"]
        params: dict[str, Any] = {}
        if search:
            where.append("d.name ILIKE :search")
            params["search"] = f"%{search.strip()}%"
        if dish_type:
            where.append("d.dish_type::text = :dtype")
            params["dtype"] = dish_type
        if status == "active":
            where.append("d.is_active")
        elif status == "inactive":
            where.append("NOT d.is_active")
        if quality == "missing_recipe":
            where.append("NOT EXISTS (SELECT 1 FROM dish_ingredients di WHERE di.dish_id = d.id)")
        elif quality == "missing_price":
            where.append("EXISTS (SELECT 1 FROM dish_ingredients di WHERE di.dish_id = d.id AND NOT EXISTS (SELECT 1 FROM price_snapshots p WHERE p.ingredient_id = di.ingredient_id AND p.price_per_default_unit IS NOT NULL))")
        elif quality == "missing_nutrition":
            where.append("EXISTS (SELECT 1 FROM dish_ingredients di LEFT JOIN nutrition_facts n ON n.ingredient_id = di.ingredient_id WHERE di.dish_id = d.id AND n.id IS NULL)")
        rows = self.session.execute(
            text(
                f"""SELECT d.id, d.code, d.name, d.dish_type, d.cooking_method, d.description,
                           d.instructions, d.tags, d.is_active,
                           COALESCE(
                               jsonb_agg(jsonb_build_object(
                                   'ingredient_id', di.ingredient_id, 'name', i.name,
                                   'quantity', di.quantity, 'unit', di.unit
                               ) ORDER BY i.name) FILTER (WHERE di.ingredient_id IS NOT NULL),
                               '[]'::jsonb
                           ) AS ingredients
                    FROM dishes d
                    LEFT JOIN dish_ingredients di ON di.dish_id = d.id
                    LEFT JOIN ingredients i ON i.id = di.ingredient_id
                    WHERE {' AND '.join(where)}
                    GROUP BY d.id
                    ORDER BY d.updated_at DESC, d.name"""
            ),
            params,
        ).fetchall()
        records = []
        for row in rows:
            data = _row_dict(row)
            ingredients = data.get("ingredients") or []
            if isinstance(ingredients, str):
                ingredients = json.loads(ingredients)
            tags = data.get("tags") or []
            if isinstance(tags, str):
                tags = json.loads(tags)
            records.append({
                "id": data["id"], "code": data["code"], "name": data["name"],
                "dish_type": _enum_value(data["dish_type"]),
                "cooking_method": _enum_value(data["cooking_method"]) if data["cooking_method"] else None,
                "description": data["description"], "instructions": data["instructions"],
                "tags": ", ".join(str(tag).strip() for tag in tags if str(tag).strip()),
                "ingredients_json": _json([
                    {
                        "ingredient_id": item["ingredient_id"], "name": item["name"],
                        "quantity": float(item["quantity"]), "unit": item["unit"],
                    }
                    for item in ingredients
                ]),
                "is_active": data["is_active"],
            })
        return self._build_export_file("dishes", output_format, records)

    def save_dish(
        self,
        data: AdminDishWrite,
        actor_id: int,
        dish_id: int | None = None,
    ) -> dict[str, Any]:
        self._validate_catalog_tags(data.tags)
        ids = [i.ingredient_id for i in data.ingredients]
        if len(ids) != len(set(ids)):
            raise ValidationAppError("Mỗi nguyên liệu chỉ được xuất hiện một lần trong công thức")
        if data.is_active and not ids:
            raise ValidationAppError("Món đang hoạt động phải có ít nhất một nguyên liệu")
        duplicate = self.session.execute(
            text(
                """SELECT id FROM dishes
                   WHERE LOWER(REGEXP_REPLACE(BTRIM(name), '\\s+', ' ', 'g')) = :name
                     AND (:id IS NULL OR id <> :id)"""
            ),
            {"name": _normalized(data.name), "id": dish_id},
        ).first()
        if duplicate:
            raise ConflictError("Tên món đã tồn tại")
        if ids:
            ingredient_rows = self.session.execute(
                text(
                    """SELECT i.id, i.default_unit, i.is_active,
                              EXISTS (SELECT 1 FROM nutrition_facts n WHERE n.ingredient_id = i.id) AS has_nutrition,
                              EXISTS (SELECT 1 FROM price_snapshots p
                                      WHERE p.ingredient_id = i.id
                                        AND p.price_per_default_unit IS NOT NULL) AS has_price
                       FROM ingredients i WHERE i.id = ANY(:ids)"""
                ),
                {"ids": ids},
            ).fetchall()
            if len(ingredient_rows) != len(ids):
                raise ValidationAppError("Công thức chứa nguyên liệu không tồn tại")
            if data.is_active:
                by_id = {row.id: row for row in ingredient_rows}
                problems: list[str] = []
                for item in data.ingredients:
                    ingredient = by_id[item.ingredient_id]
                    if not ingredient.is_active:
                        problems.append(f"nguyên liệu id={item.ingredient_id} đang inactive")
                    if not ingredient.has_nutrition:
                        problems.append(f"nguyên liệu id={item.ingredient_id} thiếu dinh dưỡng")
                    if not ingredient.has_price:
                        problems.append(f"nguyên liệu id={item.ingredient_id} thiếu giá chuẩn hoá")
                    if item.unit != ingredient.default_unit:
                        problems.append(
                            f"đơn vị '{item.unit}' của nguyên liệu id={item.ingredient_id} không khớp default_unit '{ingredient.default_unit}'"
                        )
                if problems:
                    raise ValidationAppError("Không thể kích hoạt dish: " + "; ".join(problems))
        before = self._dish_from_db(dish_id) if dish_id else None
        payload = {
            "name": data.name,
            "dtype": data.dish_type.value,
            "method": data.cooking_method.value if data.cooking_method else None,
            "description": data.description,
            "instructions": data.instructions,
            "tags": _json(data.tags),
            "active": data.is_active,
        }
        if dish_id is None:
            row = self.session.execute(
                text(
                    """INSERT INTO dishes
                       (name, dish_type, cooking_method, description, instructions, tags, is_active)
                       VALUES (:name, CAST(:dtype AS dish_type), CAST(:method AS cooking_method),
                               :description, :instructions, CAST(:tags AS jsonb), :active)
                       RETURNING id"""
                ),
                payload,
            ).first()
            dish_id = row.id
            action = "create"
        else:
            self.session.execute(
                text(
                    """UPDATE dishes SET name = :name, dish_type = CAST(:dtype AS dish_type),
                           cooking_method = CAST(:method AS cooking_method), description = :description,
                           instructions = :instructions, tags = CAST(:tags AS jsonb), is_active = :active
                       WHERE id = :id"""
                ),
                {**payload, "id": dish_id},
            )
            self.session.execute(text("DELETE FROM dish_ingredients WHERE dish_id = :id"), {"id": dish_id})
            action = "update"
        for item in data.ingredients:
            self.session.execute(
                text(
                    """INSERT INTO dish_ingredients (dish_id, ingredient_id, quantity, unit)
                       VALUES (:dish_id, :ingredient_id, :quantity, :unit)"""
                ),
                {"dish_id": dish_id, **item.model_dump()},
            )
        after = self._dish_from_db(dish_id)
        self._audit(actor_id, action, "dish", dish_id, before, after)
        self.session.commit()
        return after

    def set_dish_active(self, dish_id: int, active: bool, actor_id: int) -> dict[str, Any]:
        before = self._dish_from_db(dish_id)
        if active:
            ready = self.session.execute(
                text(
                    """SELECT ingredient_count, has_complete_nutrition, has_complete_price,
                              all_ingredients_active
                       FROM v_dishes_full WHERE id = :id"""
                ),
                {"id": dish_id},
            ).first()
            if not ready or not (
                ready.ingredient_count > 0
                and ready.has_complete_nutrition
                and ready.has_complete_price
                and ready.all_ingredients_active
            ):
                raise ValidationAppError(
                    "Không thể kích hoạt dish thiếu recipe, nutrition, price hoặc chứa nguyên liệu inactive"
                )
        self.session.execute(text("UPDATE dishes SET is_active = :active WHERE id = :id"), {"active": active, "id": dish_id})
        after = self._dish_from_db(dish_id)
        self._audit(actor_id, "restore" if active else "deactivate", "dish", dish_id, before, after)
        self.session.commit()
        return after

    def delete_dish(self, dish_id: int, actor_id: int) -> None:
        before = self._dish_from_db(dish_id)
        self.session.execute(text("DELETE FROM dishes WHERE id = :id"), {"id": dish_id})
        self._audit(actor_id, "delete", "dish", dish_id, before, None)
        self.session.commit()

    # ------------------------------------------------------------------ quality
    def _quality_cte(self) -> str:
        return f"""WITH issues AS (
            SELECT 'ingredient'::text entity_type, i.id entity_id, i.name entity_name,
                   'missing_price'::text code, 'error'::text severity,
                   'Thiếu giá chuẩn hóa'::text title,
                   'Chưa có giá quy đổi theo đơn vị mặc định.'::text detail, i.updated_at
            FROM ingredients i WHERE NOT EXISTS (
                SELECT 1 FROM price_snapshots p WHERE p.ingredient_id = i.id
                  AND p.price_per_default_unit IS NOT NULL
            )
            UNION ALL
            SELECT 'ingredient', i.id, i.name, 'missing_nutrition', 'error',
                   'Thiếu dinh dưỡng', 'Chưa có dữ liệu dinh dưỡng trên 100g.', i.updated_at
            FROM ingredients i WHERE NOT EXISTS (
                SELECT 1 FROM nutrition_facts n WHERE n.ingredient_id = i.id
            )
            UNION ALL
            SELECT 'ingredient', i.id, i.name, 'missing_conversion', 'warning',
                   'Cần kiểm tra quy đổi',
                   CASE
                       WHEN LOWER(BTRIM(i.default_unit)) IN ('ml', 'milliliter', 'milliliters')
                           THEN 'Dầu hoặc sữa đang dùng hệ số mặc định 1 g/ml; cần kiểm tra mật độ.'
                       ELSE 'Đơn vị không phải gram nhưng hệ số quy đổi đang bằng 1.'
                   END, i.updated_at
            FROM ingredients i WHERE {MISSING_CONVERSION_SQL}
            UNION ALL
            SELECT 'dish', d.id, d.name, 'missing_recipe', 'error',
                   'Thiếu công thức', 'Món chưa có nguyên liệu trong công thức.', d.updated_at
            FROM dishes d WHERE NOT EXISTS (SELECT 1 FROM dish_ingredients di WHERE di.dish_id = d.id)
            UNION ALL
            SELECT 'dish', d.id, d.name, 'missing_price', 'error',
                   'Không tính được chi phí', 'Ít nhất một nguyên liệu trong món đang thiếu giá.', d.updated_at
            FROM dishes d WHERE EXISTS (
                SELECT 1 FROM dish_ingredients di WHERE di.dish_id = d.id
                  AND NOT EXISTS (SELECT 1 FROM price_snapshots p
                      WHERE p.ingredient_id = di.ingredient_id
                        AND p.price_per_default_unit IS NOT NULL)
            )
            UNION ALL
            SELECT 'dish', d.id, d.name, 'missing_nutrition', 'error',
                   'Không tính đủ dinh dưỡng', 'Ít nhất một nguyên liệu trong món đang thiếu dinh dưỡng.', d.updated_at
            FROM dishes d WHERE EXISTS (
                SELECT 1 FROM dish_ingredients di
                LEFT JOIN nutrition_facts n ON n.ingredient_id = di.ingredient_id
                WHERE di.dish_id = d.id AND n.id IS NULL
            )
            UNION ALL
            SELECT src.entity_type, src.id, src.name, 'duplicate_name', 'warning',
                   'Tên có khả năng trùng', 'Tên chuẩn hóa trùng với một bản ghi khác.', src.updated_at
            FROM (
                SELECT 'ingredient'::text entity_type, id, name, updated_at,
                       LOWER(REGEXP_REPLACE(BTRIM(name), '\\s+', ' ', 'g')) normalized FROM ingredients
                UNION ALL SELECT 'dish', id, name, updated_at,
                       LOWER(REGEXP_REPLACE(BTRIM(name), '\\s+', ' ', 'g')) FROM dishes
            ) src JOIN (
                SELECT entity_type, normalized FROM (
                    SELECT 'ingredient'::text entity_type,
                           LOWER(REGEXP_REPLACE(BTRIM(name), '\\s+', ' ', 'g')) normalized
                    FROM ingredients
                    UNION ALL
                    SELECT 'dish', LOWER(REGEXP_REPLACE(BTRIM(name), '\\s+', ' ', 'g'))
                    FROM dishes
                ) all_names GROUP BY entity_type, normalized HAVING COUNT(*) > 1
            ) duplicates USING (entity_type, normalized)
        )"""

    def list_quality_issues(
        self,
        entity_type: str | None,
        severity: str | None,
        code: str | None,
        search: str | None,
        limit: int,
        offset: int,
    ) -> dict[str, Any]:
        where = ["1=1"]
        params: dict[str, Any] = {"limit": limit, "offset": offset}
        if entity_type:
            where.append("entity_type = :entity_type")
            params["entity_type"] = entity_type
        if severity:
            where.append("severity = :severity")
            params["severity"] = severity
        if code:
            where.append("code = :code")
            params["code"] = code
        if search:
            where.append("entity_name ILIKE :search")
            params["search"] = f"%{search.strip()}%"
        clause = " AND ".join(where)
        cte = self._quality_cte()
        total = self.session.execute(text(cte + f" SELECT COUNT(*) FROM issues WHERE {clause}"), params).scalar_one()
        rows = self.session.execute(
            text(cte + f" SELECT * FROM issues WHERE {clause} ORDER BY CASE severity WHEN 'error' THEN 0 ELSE 1 END, updated_at DESC LIMIT :limit OFFSET :offset"),
            params,
        ).fetchall()
        return {"items": [_row_dict(r) for r in rows], "total": total, "limit": limit, "offset": offset}

    # ------------------------------------------------------------------ imports
    def _build_export_file(
        self,
        entity_type: str,
        output_format: str,
        records: list[dict[str, Any]],
    ) -> tuple[bytes, str, str]:
        definitions = {
            "ingredients": {
                "headers": [
                    "id", "code", "name", "food_group", "default_unit", "grams_per_unit",
                    "tags", "calories", "protein_g", "carbs_g", "fat_g", "fiber_g", "price",
                    "price_unit", "price_per_default_unit", "source", "is_active",
                ],
                "notes": [
                    "File dùng đúng cấu trúc import nguyên liệu và có thể chỉnh sửa rồi import lại.",
                    "Giá là số sạch, không có ký hiệu tiền hoặc dấu phân cách.",
                ],
            },
            "dishes": {
                "headers": [
                    "id", "code", "name", "dish_type", "cooking_method", "description",
                    "instructions", "tags", "ingredients_json", "is_active",
                ],
                "notes": [
                    "File dùng đúng cấu trúc import món ăn (món thành phần).",
                    "Import nguyên liệu trước, sau đó mới import món ăn để các thành phần được đối chiếu.",
                ],
            },
        }
        if entity_type not in definitions or output_format not in {"csv", "xlsx"}:
            raise ValidationAppError("Không hỗ trợ loại file export này")

        definition = definitions[entity_type]
        headers = definition["headers"]
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
        filename = f"smart-menu-{entity_type}-export-{timestamp}.{output_format}"
        if output_format == "csv":
            stream = io.StringIO(newline="")
            writer = csv.DictWriter(stream, fieldnames=headers, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(records)
            return stream.getvalue().encode("utf-8-sig"), "text/csv; charset=utf-8", filename

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as exc:  # pragma: no cover - kiểm tra ở môi trường đóng gói
            raise ValidationAppError("Backend chưa cài thư viện tạo XLSX") from exc

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Dữ liệu"
        sheet.append(headers)
        for record in records:
            sheet.append([record.get(header) for header in headers])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(records) + 1)}"
        header_fill = PatternFill("solid", fgColor="047857")
        for index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=index)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            sheet.column_dimensions[get_column_letter(index)].width = min(max(len(header) + 4, 14), 28)

        guide = workbook.create_sheet("Hướng dẫn")
        guide.append(["Hướng dẫn export"])
        guide["A1"].font = Font(bold=True, color="FFFFFF")
        guide["A1"].fill = header_fill
        guide.append([f"Số bản ghi: {len(records)}"])
        for note in definition["notes"]:
            guide.append([note])
        guide.column_dimensions["A"].width = 110

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename

    def import_template(self, entity_type: str, output_format: str) -> tuple[bytes, str, str]:
        templates = {
            "ingredients": {
                "headers": [
                    "id", "code", "name", "food_group", "default_unit", "grams_per_unit",
                    "tags", "calories", "protein_g", "carbs_g", "fat_g", "fiber_g", "price",
                    "price_unit", "price_per_default_unit", "source", "is_active",
                ],
                "notes": [
                    "id: để trống khi tạo mới; điền ID hiện có khi muốn cập nhật đúng bản ghi.",
                    "code: mã riêng, duy nhất; để trống sẽ không thay đổi mã hiện có khi replace.",
                    "food_group: protein, vegetable, grain, dairy, fat, fruit hoặc other.",
                    "tags: các thẻ nguyên liệu, cách nhau bằng dấu phẩy; thẻ mới sẽ được tự tạo khi commit.",
                    "is_active: true/false; các cột dinh dưỡng và giá có thể để trống.",
                ],
            },
            "dishes": {
                "headers": [
                    "id", "code", "name", "dish_type", "cooking_method", "description",
                    "instructions", "tags", "ingredients_json", "is_active",
                ],
                "notes": [
                    "id: để trống khi tạo mới; điền ID hiện có khi muốn cập nhật đúng bản ghi.",
                    "code: mã riêng, duy nhất; để trống sẽ không thay đổi mã hiện có khi replace.",
                    "dish_type: staple, savory, soup, vegetable_side, side hoặc breakfast.",
                    "tags: các thẻ món ăn, cách nhau bằng dấu phẩy; thẻ mới sẽ được tự tạo khi commit.",
                    "ingredients_json: mảng JSON với ingredient_id hoặc name, quantity và unit.",
                    "Import nguyên liệu trước, sau đó mới import món ăn để các thành phần được đối chiếu.",
                ],
            },
        }
        if entity_type not in templates or output_format not in {"csv", "xlsx"}:
            raise ValidationAppError("Không hỗ trợ loại file mẫu này")

        headers = templates[entity_type]["headers"]
        filename = f"smart-menu-{entity_type}-template.{output_format}"
        if output_format == "csv":
            stream = io.StringIO(newline="")
            writer = csv.DictWriter(stream, fieldnames=headers)
            writer.writeheader()
            return stream.getvalue().encode("utf-8-sig"), "text/csv; charset=utf-8", filename

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as exc:  # pragma: no cover - kiểm tra ở môi trường đóng gói
            raise ValidationAppError("Backend chưa cài thư viện tạo XLSX") from exc

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Dữ liệu"
        sheet.append(headers)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        header_fill = PatternFill("solid", fgColor="047857")
        for index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=index)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            sheet.column_dimensions[get_column_letter(index)].width = min(max(len(header) + 4, 14), 28)

        guide = workbook.create_sheet("Hướng dẫn")
        guide.append(["Hướng dẫn import"])
        guide["A1"].font = Font(bold=True, color="FFFFFF")
        guide["A1"].fill = header_fill
        for note in templates[entity_type]["notes"]:
            guide.append([note])
        guide.column_dimensions["A"].width = 110

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename

    def _find_import_match(
        self,
        entity_type: str,
        record_id: int | None,
        code: str | None,
        name: str,
    ) -> tuple[dict[str, Any] | None, str | None]:
        table = "ingredients" if entity_type == "ingredients" else "dishes"
        by_id = None
        if record_id is not None:
            by_id = self.session.execute(
                text(f"SELECT id, code, name FROM {table} WHERE id = :id"), {"id": record_id}
            ).first()
            if not by_id:
                raise ValueError(f"Không tìm thấy {table[:-1]} có id={record_id}")

        by_code = None
        if code is not None:
            by_code = self.session.execute(
                text(f"SELECT id, code, name FROM {table} WHERE code = :code"), {"code": code}
            ).first()
        by_name = self.session.execute(
            text(f"SELECT id, code, name FROM {table} WHERE LOWER(REGEXP_REPLACE(BTRIM(name), '\\s+', ' ', 'g')) = :name"),
            {"name": _normalized(name)},
        ).first()

        candidates = [_row_dict(item) for item in (by_id, by_code, by_name) if item]
        if not candidates:
            return None, None
        ids = {item["id"] for item in candidates}
        if len(ids) > 1:
            raise ValueError("id, code và tên đang trỏ tới các bản ghi khác nhau")
        if by_id:
            return _row_dict(by_id), "id"
        if by_code:
            return _row_dict(by_code), "code"
        return _row_dict(by_name), "name"

    def _parse_file(self, filename: str, content: bytes) -> list[dict[str, Any]]:
        suffix = Path(filename).suffix.casefold()
        if suffix == ".csv":
            try:
                decoded = content.decode("utf-8-sig")
            except UnicodeDecodeError as exc:
                raise ValidationAppError("File CSV phải dùng mã hóa UTF-8") from exc
            try:
                dialect = csv.Sniffer().sniff(decoded[:4096], delimiters=",;\t")
            except csv.Error:
                dialect = csv.excel
            return [dict(row) for row in csv.DictReader(io.StringIO(decoded), dialect=dialect)]
        if suffix in {".xlsx", ".xlsm"}:
            try:
                from openpyxl import load_workbook
            except ImportError as exc:  # pragma: no cover - kiểm tra ở môi trường đóng gói
                raise ValidationAppError("Backend chưa cài thư viện đọc XLSX") from exc
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = [str(v).strip() if v is not None else "" for v in next(rows, [])]
            return [dict(zip(headers, row)) for row in rows if any(v is not None and str(v).strip() for v in row)]
        raise ValidationAppError("Chỉ hỗ trợ file CSV hoặc XLSX")

    def _parse_dish_ingredients(self, raw_ingredients: Any, dish_name: str) -> list[dict[str, Any]]:
        if not isinstance(raw_ingredients, list):
            raise ValueError("ingredients_json phải là một JSON array")
        parsed: list[dict[str, Any]] = []
        seen_ingredients: set[int] = set()
        for position, item in enumerate(raw_ingredients, start=1):
            if not isinstance(item, dict):
                raise ValueError(f"Thành phần #{position} phải là một object JSON")
            ingredient_id = _as_optional_id(item.get("ingredient_id"))
            ingredient_name = " ".join(str(item.get("name") or "").split())
            by_id = None
            if ingredient_id is not None:
                by_id = self.session.execute(
                    text("SELECT id, name, default_unit FROM ingredients WHERE id=:id"), {"id": ingredient_id}
                ).first()
                if not by_id:
                    raise ValueError(f"Không tìm thấy nguyên liệu id={ingredient_id} trong món {dish_name}")
            by_name = None
            if ingredient_name:
                by_name = self.session.execute(
                    text("""SELECT id, name, default_unit FROM ingredients
                            WHERE LOWER(REGEXP_REPLACE(BTRIM(name), '\\s+', ' ', 'g'))=:name"""),
                    {"name": _normalized(ingredient_name)},
                ).first()
                if not by_name:
                    raise ValueError(f"Không tìm thấy nguyên liệu '{ingredient_name}' trong món {dish_name}")
            if not by_id and not by_name:
                raise ValueError(f"Thành phần #{position} cần ingredient_id hoặc name")
            if by_id and by_name and by_id.id != by_name.id:
                raise ValueError(f"ingredient_id và name của thành phần #{position} trỏ tới hai nguyên liệu khác nhau")
            resolved_id = (by_id or by_name).id
            if resolved_id in seen_ingredients:
                raise ValueError(f"Nguyên liệu bị lặp trong món {dish_name}")
            quantity = _as_float(item.get("quantity"), "quantity", position)
            if quantity is None or quantity <= 0:
                raise ValueError(f"Định lượng thành phần #{position} phải lớn hơn 0")
            unit = str(item.get("unit") or "").strip()
            if not unit:
                raise ValueError(f"Thành phần #{position} thiếu unit")
            expected_unit = str((by_id or by_name).default_unit).strip()
            if unit.casefold() != expected_unit.casefold():
                raise ValueError(
                    f"Đơn vị '{unit}' của thành phần #{position} không khớp đơn vị mặc định '{expected_unit}'"
                )
            seen_ingredients.add(resolved_id)
            parsed.append({"ingredient_id": resolved_id, "quantity": quantity, "unit": expected_unit})
        return parsed

    def preview_import(
        self,
        entity_type: str,
        filename: str,
        content: bytes,
        actor_id: int,
    ) -> dict[str, Any]:
        if entity_type not in {"ingredients", "dishes"}:
            raise ValidationAppError("Loại import không hợp lệ")
        if len(content) > 5 * 1024 * 1024:
            raise ValidationAppError("File import không được vượt quá 5 MB")
        raw_rows = self._parse_file(filename, content)
        errors: list[dict] = []
        warnings: list[dict] = []
        conflicts: list[dict] = []
        valid: list[dict] = []
        seen_names: set[str] = set()
        seen_codes: set[str] = set()
        seen_target_ids: set[int] = set()
        for index, raw in enumerate(raw_rows, start=2):
            try:
                name = " ".join(str(raw.get("name") or "").split())
                if not name:
                    raise ValueError("Thiếu cột name")
                key = _normalized(name)
                if key in seen_names:
                    raise ValueError("Tên bị lặp trong cùng file")
                seen_names.add(key)
                record_id = _as_optional_id(raw.get("id"))
                code = _normalized_code(raw.get("code"))
                if code is not None:
                    if code in seen_codes:
                        raise ValueError("code bị lặp trong cùng file")
                    seen_codes.add(code)
                existing, match_by = self._find_import_match(entity_type, record_id, code, name)
                if existing:
                    if existing["id"] in seen_target_ids:
                        raise ValueError("Nhiều dòng đang trỏ tới cùng một bản ghi")
                    seen_target_ids.add(existing["id"])
                    conflicts.append({
                        "row": index,
                        "match_by": match_by,
                        "incoming": {"id": record_id, "code": code, "name": name},
                        "existing": existing,
                    })
                if entity_type == "ingredients":
                    group = str(raw.get("food_group") or "").strip()
                    if group not in {v.value for v in FoodGroup}:
                        raise ValueError("food_group không hợp lệ")
                    nutrition_fields = ["calories", "protein_g", "carbs_g", "fat_g", "fiber_g"]
                    has_nutrition = any(str(raw.get(k) or "").strip() for k in nutrition_fields)
                    nutrition = None
                    if has_nutrition:
                        nutrition = {k: _as_float(raw.get(k), k, index, default=0) for k in nutrition_fields}
                    has_price = str(raw.get("price") or "").strip() != ""
                    price = None
                    if has_price:
                        price = {
                            "price": _as_money(raw.get("price"), "price", index, default=0),
                            "unit": str(raw.get("price_unit") or raw.get("unit") or "kg").strip(),
                            "price_per_default_unit": _as_money(
                                raw.get("price_per_default_unit"), "price_per_default_unit", index,
                                max_fraction_digits=4,
                            ),
                            "source": str(raw.get("source") or "").strip() or None,
                        }
                        if price["price_per_default_unit"] is None:
                            raise ValueError("Có price nhưng thiếu price_per_default_unit")
                    valid.append({
                        "source_row": index,
                        "id": record_id,
                        "code": code,
                        "name": name,
                        "food_group": group,
                        "default_unit": str(raw.get("default_unit") or "g").strip(),
                        "grams_per_unit": _as_float(raw.get("grams_per_unit"), "grams_per_unit", index, default=1),
                        "tags": _normalized_tags(raw.get("tags")),
                        "is_active": _as_bool(raw.get("is_active"), True),
                        "nutrition": nutrition,
                        "price": price,
                    })
                else:
                    dtype = str(raw.get("dish_type") or "").strip()
                    if dtype not in {v.value for v in DishType}:
                        raise ValueError("dish_type không hợp lệ")
                    ingredients_raw = raw.get("ingredients_json") or raw.get("ingredients") or "[]"
                    ingredients = json.loads(ingredients_raw) if isinstance(ingredients_raw, str) else ingredients_raw
                    ingredients = self._parse_dish_ingredients(ingredients, name)
                    valid.append({
                        "source_row": index,
                        "id": record_id,
                        "code": code,
                        "name": name,
                        "dish_type": dtype,
                        "cooking_method": str(raw.get("cooking_method") or "").strip() or None,
                        "description": str(raw.get("description") or "").strip() or None,
                        "instructions": str(raw.get("instructions") or "").strip() or None,
                        "tags": _normalized_tags(raw.get("tags")),
                        "is_active": _as_bool(raw.get("is_active"), True),
                        "ingredients": ingredients,
                    })
            except (ValueError, json.JSONDecodeError) as exc:
                errors.append({"row": index, "message": str(exc)})
        status = "validated" if not errors else "invalid"
        row = self.session.execute(
            text(
                """INSERT INTO import_jobs
                   (entity_type, filename, status, payload, errors, warnings, conflicts,
                    total_rows, valid_rows, error_count, created_by)
                   VALUES (:entity_type, :filename, :status, CAST(:payload AS jsonb),
                           CAST(:errors AS jsonb), CAST(:warnings AS jsonb), CAST(:conflicts AS jsonb),
                           :total, :valid, :error_count, :actor) RETURNING id"""
            ),
            {
                "entity_type": entity_type, "filename": filename, "status": status,
                "payload": _json(valid), "errors": _json(errors), "warnings": _json(warnings),
                "conflicts": _json(conflicts),
                "total": len(raw_rows), "valid": len(valid), "error_count": len(errors),
                "actor": actor_id,
            },
        ).first()
        self.session.commit()
        return {
            "job_id": row.id,
            "entity_type": entity_type,
            "filename": filename,
            "total_rows": len(raw_rows),
            "valid_rows": len(valid),
            "errors": errors,
            "warnings": warnings,
            "conflicts": conflicts,
            "preview": valid[:20],
            "can_commit": not errors and bool(valid),
        }

    def commit_import(self, job_id: int, replace_rows: list[int], actor_id: int) -> dict[str, Any]:
        job = self.session.execute(text("SELECT * FROM import_jobs WHERE id = :id FOR UPDATE"), {"id": job_id}).first()
        if not job:
            raise NotFoundError("Không tìm thấy phiên import")
        job_data = _row_dict(job)
        if job_data["created_by"] != actor_id:
            raise ValidationAppError("Chỉ người tạo preview mới được commit phiên import này")
        if job_data["status"] != "validated":
            raise ValidationAppError("Phiên import chưa hợp lệ hoặc đã được commit")
        rows = job_data["payload"] or []
        conflicts_by_row = {int(item["row"]): item for item in (job_data.get("conflicts") or [])}
        replace_set = set(replace_rows)
        unknown_rows = replace_set - set(conflicts_by_row)
        if unknown_rows:
            raise ValidationAppError("Có lựa chọn replace không thuộc phiên preview này")
        created = 0
        updated = 0
        skipped = 0
        try:
            for raw in rows:
                source_row = int(raw.get("source_row") or 0)
                conflict = conflicts_by_row.get(source_row)
                existing, _ = self._find_import_match(
                    job_data["entity_type"], raw.get("id"), raw.get("code"), raw["name"]
                )
                if conflict and (not existing or existing["id"] != conflict["existing"]["id"]):
                    raise ValidationAppError("Dữ liệu đã thay đổi sau preview; hãy kiểm tra file lại")
                if not conflict and existing:
                    raise ValidationAppError("Dữ liệu đã thay đổi sau preview; hãy kiểm tra file lại")
                if conflict and source_row not in replace_set:
                    skipped += 1
                    continue
                next_code = raw.get("code") if raw.get("code") is not None else (existing or {}).get("code")
                if job_data["entity_type"] == "ingredients":
                    tags = _normalized_tags(raw.get("tags"))
                    self._ensure_catalog_tags(tags, "ingredient")
                    if existing:
                        ingredient_id = existing["id"]
                        self.session.execute(
                            text("""UPDATE ingredients SET code=:code, name=:name, food_group=CAST(:group AS food_group),
                                   default_unit=:unit, grams_per_unit=:grams, tags=CAST(:tags AS jsonb),
                                   is_active=:active WHERE id=:id"""),
                            {"id": ingredient_id, "code": next_code, "name": raw["name"], "group": raw["food_group"],
                             "unit": raw["default_unit"], "grams": raw["grams_per_unit"], "tags": _json(tags),
                             "active": raw["is_active"]},
                        )
                        updated += 1
                    else:
                        inserted = self.session.execute(
                            text("""INSERT INTO ingredients
                                   (code, name, food_group, default_unit, grams_per_unit, tags, is_active)
                                   VALUES (:code, :name, CAST(:group AS food_group), :unit, :grams,
                                           CAST(:tags AS jsonb), :active) RETURNING id"""),
                            {"code": next_code, "name": raw["name"], "group": raw["food_group"], "unit": raw["default_unit"],
                             "grams": raw["grams_per_unit"], "tags": _json(tags), "active": raw["is_active"]},
                        ).first()
                        ingredient_id = inserted.id
                        created += 1
                    if raw.get("nutrition") is not None:
                        self.session.execute(
                            text("""INSERT INTO nutrition_facts
                                   (ingredient_id, calories, protein_g, carbs_g, fat_g, fiber_g)
                                   VALUES (:ingredient_id, :calories, :protein_g, :carbs_g, :fat_g, :fiber_g)
                                   ON CONFLICT (ingredient_id) DO UPDATE SET calories=EXCLUDED.calories,
                                   protein_g=EXCLUDED.protein_g, carbs_g=EXCLUDED.carbs_g,
                                   fat_g=EXCLUDED.fat_g, fiber_g=EXCLUDED.fiber_g"""),
                            {"ingredient_id": ingredient_id, **raw["nutrition"]},
                        )
                    if raw.get("price") is not None:
                        self.session.execute(
                            text("""INSERT INTO price_snapshots
                                   (ingredient_id, price, unit, price_per_default_unit, source)
                                   VALUES (:ingredient_id, :price, :unit, :price_per_default_unit, :source)"""),
                            {"ingredient_id": ingredient_id, **raw["price"]},
                        )
                else:
                    tags = _normalized_tags(raw.get("tags"))
                    self._ensure_catalog_tags(tags, "dish")
                    ingredient_payload = []
                    for item in raw.get("ingredients", []):
                        ingredient_id = _as_optional_id(item.get("ingredient_id"))
                        found = self.session.execute(
                            text("SELECT id FROM ingredients WHERE id=:id"), {"id": ingredient_id}
                        ).first() if ingredient_id else None
                        if not found:
                            raise ValidationAppError(f"Không tìm thấy nguyên liệu trong món {raw['name']}")
                        ingredient_payload.append({
                            "ingredient_id": found.id,
                            "quantity": float(item.get("quantity") or 0),
                            "unit": str(item.get("unit") or "").strip(),
                        })
                    if existing:
                        dish_id = existing["id"]
                        self.session.execute(
                            text("""UPDATE dishes SET code=:code, name=:name, dish_type=CAST(:dtype AS dish_type),
                                   cooking_method=CAST(:method AS cooking_method), description=:description,
                                   instructions=:instructions, tags=CAST(:tags AS jsonb), is_active=:active
                                   WHERE id=:id"""),
                            {"id": dish_id, "code": next_code, "name": raw["name"], "dtype": raw["dish_type"],
                             "method": raw.get("cooking_method"), "description": raw.get("description"),
                             "instructions": raw.get("instructions"), "tags": _json(tags),
                             "active": raw.get("is_active", True)},
                        )
                        self.session.execute(text("DELETE FROM dish_ingredients WHERE dish_id=:id"), {"id": dish_id})
                        updated += 1
                    else:
                        inserted = self.session.execute(
                            text("""INSERT INTO dishes
                                   (code, name, dish_type, cooking_method, description, instructions, tags, is_active)
                                   VALUES (:code, :name, CAST(:dtype AS dish_type), CAST(:method AS cooking_method),
                                   :description, :instructions, CAST(:tags AS jsonb), :active) RETURNING id"""),
                            {"code": next_code, "name": raw["name"], "dtype": raw["dish_type"], "method": raw.get("cooking_method"),
                             "description": raw.get("description"), "instructions": raw.get("instructions"),
                             "tags": _json(tags), "active": raw.get("is_active", True)},
                        ).first()
                        dish_id = inserted.id
                        created += 1
                    for item in ingredient_payload:
                        if item["quantity"] <= 0:
                            raise ValidationAppError(f"Định lượng trong món {raw['name']} phải lớn hơn 0")
                        if not item["unit"]:
                            raise ValidationAppError(f"Thiếu đơn vị trong món {raw['name']}")
                        self.session.execute(
                            text("""INSERT INTO dish_ingredients (dish_id, ingredient_id, quantity, unit)
                                   VALUES (:dish_id, :ingredient_id, :quantity, :unit)"""),
                            {"dish_id": dish_id, **item},
                        )
            summary = {"created": created, "updated": updated, "skipped": skipped}
            self.session.execute(
                text("UPDATE import_jobs SET status='committed', completed_at=NOW() WHERE id=:id"),
                {"id": job_id},
            )
            self._audit(actor_id, "import", job_data["entity_type"], job_id, after=summary)
            self.session.commit()
            return {"job_id": job_id, "status": "committed", **summary}
        except Exception:
            self.session.rollback()
            raise

    def list_import_jobs(self, limit: int, offset: int) -> dict[str, Any]:
        total = self.session.execute(text("SELECT COUNT(*) FROM import_jobs")).scalar_one()
        rows = self.session.execute(
            text("""SELECT id, entity_type, filename, status, total_rows, valid_rows,
                          error_count, created_by, created_at, completed_at
                   FROM import_jobs ORDER BY created_at DESC LIMIT :limit OFFSET :offset"""),
            {"limit": limit, "offset": offset},
        ).fetchall()
        return {"items": [_row_dict(r) for r in rows], "total": total, "limit": limit, "offset": offset}
