from __future__ import annotations

import csv
import io
import json
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from sqlalchemy import text

from app.core.exceptions import NotFoundError, ValidationAppError
from app.modules.admin.normalization import (
    as_bool as _as_bool,
    as_float as _as_float,
    as_money as _as_money,
    as_optional_id as _as_optional_id,
    json_value as _json,
    normalized as _normalized,
    normalized_code as _normalized_code,
    normalized_tags as _normalized_tags,
    row_dict as _row_dict,
)
from app.shared.enums import DishType, FoodGroup


class AdminImportMixin:
    def _build_export_file(
        self,
        entity_type: str,
        output_format: str,
        records: list[dict[str, Any]],
    ) -> tuple[bytes, str, str]:
        definitions = {
            "ingredients": {
                "headers": [
                    "id", "code", "name", "food_group", "default_unit", "grams_per_unit",
                    "tags", "calories", "protein_g", "carbs_g", "fat_g", "fiber_g", "price",
                    "price_unit", "price_per_default_unit", "source", "is_active",
                    "purchase_mode", "purchase_increment", "room_shelf_life_days",
                    "fridge_shelf_life_days", "freezer_shelf_life_days",
                    "shelf_life_source", "shelf_life_reviewed_at",
                ],
                "notes": [
                    "File dùng đúng cấu trúc import nguyên liệu và có thể chỉnh sửa rồi import lại.",
                    "Giá là số sạch, không có ký hiệu tiền hoặc dấu phân cách.",
                ],
            },
            "dishes": {
                "headers": [
                    "id", "code", "name", "dish_type", "cooking_method", "description",
                    "instructions", "tags", "ingredients_json", "is_active",
                ],
                "notes": [
                    "File dùng đúng cấu trúc import món ăn (món thành phần).",
                    "Import nguyên liệu trước, sau đó mới import món ăn để các thành phần được đối chiếu.",
                ],
            },
        }
        if entity_type not in definitions or output_format not in {"csv", "xlsx"}:
            raise ValidationAppError("Không hỗ trợ loại file export này")

        definition = definitions[entity_type]
        headers = definition["headers"]
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
        filename = f"smart-menu-{entity_type}-export-{timestamp}.{output_format}"
        if output_format == "csv":
            stream = io.StringIO(newline="")
            writer = csv.DictWriter(stream, fieldnames=headers, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(records)
            return stream.getvalue().encode("utf-8-sig"), "text/csv; charset=utf-8", filename

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as exc:  # pragma: no cover - kiểm tra ở môi trường đóng gói
            raise ValidationAppError("Backend chưa cài thư viện tạo XLSX") from exc

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Dữ liệu"
        sheet.append(headers)
        for record in records:
            sheet.append([record.get(header) for header in headers])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(records) + 1)}"
        header_fill = PatternFill("solid", fgColor="047857")
        for index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=index)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            sheet.column_dimensions[get_column_letter(index)].width = min(max(len(header) + 4, 14), 28)

        guide = workbook.create_sheet("Hướng dẫn")
        guide.append(["Hướng dẫn export"])
        guide["A1"].font = Font(bold=True, color="FFFFFF")
        guide["A1"].fill = header_fill
        guide.append([f"Số bản ghi: {len(records)}"])
        for note in definition["notes"]:
            guide.append([note])
        guide.column_dimensions["A"].width = 110

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename

    def import_template(self, entity_type: str, output_format: str) -> tuple[bytes, str, str]:
        templates = {
            "ingredients": {
                "headers": [
                    "id", "code", "name", "food_group", "default_unit", "grams_per_unit",
                    "tags", "calories", "protein_g", "carbs_g", "fat_g", "fiber_g", "price",
                    "price_unit", "price_per_default_unit", "source", "is_active",
                    "purchase_mode", "purchase_increment", "room_shelf_life_days",
                    "fridge_shelf_life_days", "freezer_shelf_life_days",
                    "shelf_life_source", "shelf_life_reviewed_at",
                ],
                "notes": [
                    "id: để trống khi tạo mới; điền ID hiện có khi muốn cập nhật đúng bản ghi.",
                    "code: mã riêng, duy nhất; để trống sẽ không thay đổi mã hiện có khi replace.",
                    "food_group: protein, vegetable, grain, dairy, fat, fruit hoặc other.",
                    "tags: các thẻ nguyên liệu, cách nhau bằng dấu phẩy; thẻ mới sẽ được tự tạo khi commit.",
                    "is_active: true/false; các cột dinh dưỡng và giá có thể để trống.",
                    "purchase_mode: regular, pantry hoặc ignored; regular mua theo bội số purchase_increment.",
                    "Shelf-life là số ngày cộng thêm sau ngày mua; có giá trị thì cần source và reviewed_at.",
                ],
            },
            "dishes": {
                "headers": [
                    "id", "code", "name", "dish_type", "cooking_method", "description",
                    "instructions", "tags", "ingredients_json", "is_active",
                ],
                "notes": [
                    "id: để trống khi tạo mới; điền ID hiện có khi muốn cập nhật đúng bản ghi.",
                    "code: mã riêng, duy nhất; để trống sẽ không thay đổi mã hiện có khi replace.",
                    "dish_type: staple, savory, soup, vegetable_side, side hoặc breakfast.",
                    "tags: các thẻ món ăn, cách nhau bằng dấu phẩy; thẻ mới sẽ được tự tạo khi commit.",
                    "ingredients_json: mảng JSON với ingredient_id hoặc name, quantity, unit, max_extra_quantity và extra_step_quantity.",
                    "Import nguyên liệu trước, sau đó mới import món ăn để các thành phần được đối chiếu.",
                ],
            },
        }
        if entity_type not in templates or output_format not in {"csv", "xlsx"}:
            raise ValidationAppError("Không hỗ trợ loại file mẫu này")

        headers = templates[entity_type]["headers"]
        filename = f"smart-menu-{entity_type}-template.{output_format}"
        if output_format == "csv":
            stream = io.StringIO(newline="")
            writer = csv.DictWriter(stream, fieldnames=headers)
            writer.writeheader()
            return stream.getvalue().encode("utf-8-sig"), "text/csv; charset=utf-8", filename

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as exc:  # pragma: no cover - kiểm tra ở môi trường đóng gói
            raise ValidationAppError("Backend chưa cài thư viện tạo XLSX") from exc

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Dữ liệu"
        sheet.append(headers)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        header_fill = PatternFill("solid", fgColor="047857")
        for index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=index)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            sheet.column_dimensions[get_column_letter(index)].width = min(max(len(header) + 4, 14), 28)

        guide = workbook.create_sheet("Hướng dẫn")
        guide.append(["Hướng dẫn import"])
        guide["A1"].font = Font(bold=True, color="FFFFFF")
        guide["A1"].fill = header_fill
        for note in templates[entity_type]["notes"]:
            guide.append([note])
        guide.column_dimensions["A"].width = 110

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename

    def _find_import_match(
        self,
        entity_type: str,
        record_id: int | None,
        code: str | None,
        name: str,
    ) -> tuple[dict[str, Any] | None, str | None]:
        table = "ingredients" if entity_type == "ingredients" else "dishes"
        by_id = None
        if record_id is not None:
            by_id = self.session.execute(
                text(f"SELECT id, code, name FROM {table} WHERE id = :id"), {"id": record_id}
            ).first()
            if not by_id:
                raise ValueError(f"Không tìm thấy {table[:-1]} có id={record_id}")

        by_code = None
        if code is not None:
            by_code = self.session.execute(
                text(f"SELECT id, code, name FROM {table} WHERE code = :code"), {"code": code}
            ).first()
        by_name = self.session.execute(
            text(f"SELECT id, code, name FROM {table} WHERE LOWER(REGEXP_REPLACE(BTRIM(name), '\\s+', ' ', 'g')) = :name"),
            {"name": _normalized(name)},
        ).first()

        candidates = [_row_dict(item) for item in (by_id, by_code, by_name) if item]
        if not candidates:
            return None, None
        ids = {item["id"] for item in candidates}
        if len(ids) > 1:
            raise ValueError("id, code và tên đang trỏ tới các bản ghi khác nhau")
        if by_id:
            return _row_dict(by_id), "id"
        if by_code:
            return _row_dict(by_code), "code"
        return _row_dict(by_name), "name"

    def _parse_file(self, filename: str, content: bytes) -> list[dict[str, Any]]:
        suffix = Path(filename).suffix.casefold()
        if suffix == ".csv":
            try:
                decoded = content.decode("utf-8-sig")
            except UnicodeDecodeError as exc:
                raise ValidationAppError("File CSV phải dùng mã hóa UTF-8") from exc
            try:
                dialect = csv.Sniffer().sniff(decoded[:4096], delimiters=",;\t")
            except csv.Error:
                dialect = csv.excel
            return [dict(row) for row in csv.DictReader(io.StringIO(decoded), dialect=dialect)]
        if suffix in {".xlsx", ".xlsm"}:
            try:
                from openpyxl import load_workbook
            except ImportError as exc:  # pragma: no cover - kiểm tra ở môi trường đóng gói
                raise ValidationAppError("Backend chưa cài thư viện đọc XLSX") from exc
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = [str(v).strip() if v is not None else "" for v in next(rows, [])]
            return [
                dict(zip(headers, row, strict=False))
                for row in rows
                if any(v is not None and str(v).strip() for v in row)
            ]
        raise ValidationAppError("Chỉ hỗ trợ file CSV hoặc XLSX")

    def _parse_dish_ingredients(self, raw_ingredients: Any, dish_name: str) -> list[dict[str, Any]]:
        if not isinstance(raw_ingredients, list):
            raise ValueError("ingredients_json phải là một JSON array")
        parsed: list[dict[str, Any]] = []
        seen_ingredients: set[int] = set()
        for position, item in enumerate(raw_ingredients, start=1):
            if not isinstance(item, dict):
                raise ValueError(f"Thành phần #{position} phải là một object JSON")
            ingredient_id = _as_optional_id(item.get("ingredient_id"))
            ingredient_name = " ".join(str(item.get("name") or "").split())
            by_id = None
            if ingredient_id is not None:
                by_id = self.session.execute(
                    text("SELECT id, name, default_unit FROM ingredients WHERE id=:id"), {"id": ingredient_id}
                ).first()
                if not by_id:
                    raise ValueError(f"Không tìm thấy nguyên liệu id={ingredient_id} trong món {dish_name}")
            by_name = None
            if ingredient_name:
                by_name = self.session.execute(
                    text("""SELECT id, name, default_unit FROM ingredients
                            WHERE LOWER(REGEXP_REPLACE(BTRIM(name), '\\s+', ' ', 'g'))=:name"""),
                    {"name": _normalized(ingredient_name)},
                ).first()
                if not by_name:
                    raise ValueError(f"Không tìm thấy nguyên liệu '{ingredient_name}' trong món {dish_name}")
            if not by_id and not by_name:
                raise ValueError(f"Thành phần #{position} cần ingredient_id hoặc name")
            if by_id and by_name and by_id.id != by_name.id:
                raise ValueError(f"ingredient_id và name của thành phần #{position} trỏ tới hai nguyên liệu khác nhau")
            resolved_id = (by_id or by_name).id
            if resolved_id in seen_ingredients:
                raise ValueError(f"Nguyên liệu bị lặp trong món {dish_name}")
            quantity = _as_float(item.get("quantity"), "quantity", position)
            if quantity is None or quantity <= 0:
                raise ValueError(f"Định lượng thành phần #{position} phải lớn hơn 0")
            unit = str(item.get("unit") or "").strip()
            if not unit:
                raise ValueError(f"Thành phần #{position} thiếu unit")
            expected_unit = str((by_id or by_name).default_unit).strip()
            if unit.casefold() != expected_unit.casefold():
                raise ValueError(
                    f"Đơn vị '{unit}' của thành phần #{position} không khớp đơn vị mặc định '{expected_unit}'"
                )
            max_extra = _as_float(
                item.get("max_extra_quantity"), "max_extra_quantity", position, default=0
            )
            step_extra = _as_float(
                item.get("extra_step_quantity"), "extra_step_quantity", position
            )
            if max_extra is None or max_extra < 0:
                raise ValueError(f"Mức tăng tối đa của thành phần #{position} không hợp lệ")
            if max_extra == 0 and step_extra is not None:
                raise ValueError(f"Thành phần #{position} cố định nhưng lại có bước tăng")
            if max_extra > 0:
                if step_extra is None or step_extra <= 0 or step_extra > max_extra:
                    raise ValueError(f"Bước tăng của thành phần #{position} không hợp lệ")
                quotient = max_extra / step_extra
                if abs(quotient - round(quotient)) > 1e-6:
                    raise ValueError(f"Mức tăng tối đa của thành phần #{position} phải chia hết cho bước tăng")
            seen_ingredients.add(resolved_id)
            parsed_item = {
                "ingredient_id": resolved_id,
                "quantity": quantity,
                "unit": expected_unit,
            }
            if "max_extra_quantity" in item or "extra_step_quantity" in item:
                parsed_item.update({
                    "max_extra_quantity": max_extra,
                    "extra_step_quantity": step_extra,
                })
            parsed.append(parsed_item)
        return parsed

    def preview_import(
        self,
        entity_type: str,
        filename: str,
        content: bytes,
        actor_id: int,
    ) -> dict[str, Any]:
        if entity_type not in {"ingredients", "dishes"}:
            raise ValidationAppError("Loại import không hợp lệ")
        if len(content) > 5 * 1024 * 1024:
            raise ValidationAppError("File import không được vượt quá 5 MB")
        raw_rows = self._parse_file(filename, content)
        errors: list[dict] = []
        warnings: list[dict] = []
        conflicts: list[dict] = []
        valid: list[dict] = []
        seen_names: set[str] = set()
        seen_codes: set[str] = set()
        seen_target_ids: set[int] = set()
        for index, raw in enumerate(raw_rows, start=2):
            try:
                name = " ".join(str(raw.get("name") or "").split())
                if not name:
                    raise ValueError("Thiếu cột name")
                key = _normalized(name)
                if key in seen_names:
                    raise ValueError("Tên bị lặp trong cùng file")
                seen_names.add(key)
                record_id = _as_optional_id(raw.get("id"))
                code = _normalized_code(raw.get("code"))
                if code is not None:
                    if code in seen_codes:
                        raise ValueError("code bị lặp trong cùng file")
                    seen_codes.add(code)
                existing, match_by = self._find_import_match(entity_type, record_id, code, name)
                if existing:
                    if existing["id"] in seen_target_ids:
                        raise ValueError("Nhiều dòng đang trỏ tới cùng một bản ghi")
                    seen_target_ids.add(existing["id"])
                    conflicts.append({
                        "row": index,
                        "match_by": match_by,
                        "incoming": {"id": record_id, "code": code, "name": name},
                        "existing": existing,
                    })
                if entity_type == "ingredients":
                    group = str(raw.get("food_group") or "").strip()
                    if group not in {v.value for v in FoodGroup}:
                        raise ValueError("food_group không hợp lệ")
                    nutrition_fields = ["calories", "protein_g", "carbs_g", "fat_g", "fiber_g"]
                    has_nutrition = any(str(raw.get(k) or "").strip() for k in nutrition_fields)
                    nutrition = None
                    if has_nutrition:
                        nutrition = {k: _as_float(raw.get(k), k, index, default=0) for k in nutrition_fields}
                    has_price = str(raw.get("price") or "").strip() != ""
                    price = None
                    if has_price:
                        price = {
                            "price": _as_money(raw.get("price"), "price", index, default=0),
                            "unit": str(raw.get("price_unit") or raw.get("unit") or "kg").strip(),
                            "price_per_default_unit": _as_money(
                                raw.get("price_per_default_unit"), "price_per_default_unit", index,
                                max_fraction_digits=4,
                            ),
                            "source": str(raw.get("source") or "").strip() or None,
                        }
                        if price["price_per_default_unit"] is None:
                            raise ValueError("Có price nhưng thiếu price_per_default_unit")
                    procurement_fields = {
                        "purchase_mode", "purchase_increment", "room_shelf_life_days",
                        "fridge_shelf_life_days", "freezer_shelf_life_days",
                        "shelf_life_source", "shelf_life_reviewed_at",
                    }
                    procurement_provided = any(field in raw for field in procurement_fields)
                    raw_mode = str(raw.get("purchase_mode") or "").strip()
                    purchase_mode = raw_mode or ("regular" if not existing else None)
                    if purchase_mode is not None and purchase_mode not in {"regular", "pantry", "ignored"}:
                        raise ValueError("purchase_mode không hợp lệ")
                    purchase_increment = _as_float(
                        raw.get("purchase_increment"), "purchase_increment", index
                    )
                    if purchase_increment is not None and purchase_increment <= 0:
                        raise ValueError("purchase_increment phải lớn hơn 0")

                    shelf_life: dict[str, int | None] = {}
                    for field in (
                        "room_shelf_life_days", "fridge_shelf_life_days", "freezer_shelf_life_days"
                    ):
                        value = _as_float(raw.get(field), field, index)
                        if value is not None and (value < 0 or value > 3650 or value != int(value)):
                            raise ValueError(f"{field} phải là số nguyên từ 0 đến 3650")
                        shelf_life[field] = int(value) if value is not None else None
                    shelf_source = str(raw.get("shelf_life_source") or "").strip() or None
                    reviewed_raw = str(raw.get("shelf_life_reviewed_at") or "").strip()
                    reviewed_at = None
                    if reviewed_raw:
                        try:
                            reviewed_at = date.fromisoformat(reviewed_raw[:10])
                        except ValueError as exc:
                            raise ValueError("shelf_life_reviewed_at phải có dạng YYYY-MM-DD") from exc
                        if reviewed_at > date.today():
                            raise ValueError("shelf_life_reviewed_at không được ở tương lai")
                    if any(value is not None for value in shelf_life.values()) and (
                        shelf_source is None or reviewed_at is None
                    ):
                        raise ValueError("Có hạn bảo quản thì phải có source và reviewed_at")
                    if purchase_mode in {"pantry", "ignored"} and (
                        purchase_increment is not None
                        or any(value is not None for value in shelf_life.values())
                        or shelf_source is not None
                        or reviewed_at is not None
                    ):
                        raise ValueError("pantry/ignored không được khai báo quy cách mua hoặc bảo quản")
                    effective_mode = purchase_mode or "regular"
                    if effective_mode == "regular" and purchase_increment is None and (
                        procurement_provided or not existing
                    ):
                        warnings.append({
                            "row": index,
                            "field": "purchase_increment",
                            "message": "Nguyên liệu regular chưa có bước mua và sẽ không planner-ready V3.",
                        })
                    if effective_mode == "regular" and not any(
                        value is not None for value in shelf_life.values()
                    ) and (procurement_provided or not existing):
                        warnings.append({
                            "row": index,
                            "field": "storage",
                            "message": "Chưa có hạn bảo quản; planner chỉ mua và dùng trong cùng ngày.",
                        })
                    valid.append({
                        "source_row": index,
                        "id": record_id,
                        "code": code,
                        "name": name,
                        "food_group": group,
                        "default_unit": str(raw.get("default_unit") or "g").strip(),
                        "grams_per_unit": _as_float(raw.get("grams_per_unit"), "grams_per_unit", index, default=1),
                        "tags": _normalized_tags(raw.get("tags")),
                        "is_active": _as_bool(raw.get("is_active"), True),
                        "nutrition": nutrition,
                        "price": price,
                        "procurement_provided": procurement_provided,
                        "purchase_mode": purchase_mode,
                        "purchase_increment": purchase_increment,
                        **shelf_life,
                        "shelf_life_source": shelf_source,
                        "shelf_life_reviewed_at": reviewed_at.isoformat() if reviewed_at else None,
                    })
                else:
                    dtype = str(raw.get("dish_type") or "").strip()
                    if dtype not in {v.value for v in DishType}:
                        raise ValueError("dish_type không hợp lệ")
                    ingredients_raw = raw.get("ingredients_json") or raw.get("ingredients") or "[]"
                    ingredients = json.loads(ingredients_raw) if isinstance(ingredients_raw, str) else ingredients_raw
                    ingredients = self._parse_dish_ingredients(ingredients, name)
                    valid.append({
                        "source_row": index,
                        "id": record_id,
                        "code": code,
                        "name": name,
                        "dish_type": dtype,
                        "cooking_method": str(raw.get("cooking_method") or "").strip() or None,
                        "description": str(raw.get("description") or "").strip() or None,
                        "instructions": str(raw.get("instructions") or "").strip() or None,
                        "tags": _normalized_tags(raw.get("tags")),
                        "is_active": _as_bool(raw.get("is_active"), True),
                        "ingredients": ingredients,
                    })
            except (ValueError, json.JSONDecodeError) as exc:
                errors.append({"row": index, "message": str(exc)})
        status = "validated" if not errors else "invalid"
        row = self.session.execute(
            text(
                """INSERT INTO import_jobs
                   (entity_type, filename, status, payload, errors, warnings, conflicts,
                    total_rows, valid_rows, error_count, created_by)
                   VALUES (:entity_type, :filename, :status, CAST(:payload AS jsonb),
                           CAST(:errors AS jsonb), CAST(:warnings AS jsonb), CAST(:conflicts AS jsonb),
                           :total, :valid, :error_count, :actor) RETURNING id"""
            ),
            {
                "entity_type": entity_type, "filename": filename, "status": status,
                "payload": _json(valid), "errors": _json(errors), "warnings": _json(warnings),
                "conflicts": _json(conflicts),
                "total": len(raw_rows), "valid": len(valid), "error_count": len(errors),
                "actor": actor_id,
            },
        ).first()
        self.session.commit()
        return {
            "job_id": row.id,
            "entity_type": entity_type,
            "filename": filename,
            "total_rows": len(raw_rows),
            "valid_rows": len(valid),
            "errors": errors,
            "warnings": warnings,
            "conflicts": conflicts,
            "preview": valid[:20],
            "can_commit": not errors and bool(valid),
        }

    def commit_import(self, job_id: int, replace_rows: list[int], actor_id: int) -> dict[str, Any]:
        job = self.session.execute(text("SELECT * FROM import_jobs WHERE id = :id FOR UPDATE"), {"id": job_id}).first()
        if not job:
            raise NotFoundError("Không tìm thấy phiên import")
        job_data = _row_dict(job)
        if job_data["created_by"] != actor_id:
            raise ValidationAppError("Chỉ người tạo preview mới được commit phiên import này")
        if job_data["status"] != "validated":
            raise ValidationAppError("Phiên import chưa hợp lệ hoặc đã được commit")
        rows = job_data["payload"] or []
        conflicts_by_row = {int(item["row"]): item for item in (job_data.get("conflicts") or [])}
        replace_set = set(replace_rows)
        unknown_rows = replace_set - set(conflicts_by_row)
        if unknown_rows:
            raise ValidationAppError("Có lựa chọn replace không thuộc phiên preview này")
        created = 0
        updated = 0
        skipped = 0
        try:
            for raw in rows:
                source_row = int(raw.get("source_row") or 0)
                conflict = conflicts_by_row.get(source_row)
                existing, _ = self._find_import_match(
                    job_data["entity_type"], raw.get("id"), raw.get("code"), raw["name"]
                )
                if conflict and (not existing or existing["id"] != conflict["existing"]["id"]):
                    raise ValidationAppError("Dữ liệu đã thay đổi sau preview; hãy kiểm tra file lại")
                if not conflict and existing:
                    raise ValidationAppError("Dữ liệu đã thay đổi sau preview; hãy kiểm tra file lại")
                if conflict and source_row not in replace_set:
                    skipped += 1
                    continue
                next_code = raw.get("code") if raw.get("code") is not None else (existing or {}).get("code")
                if job_data["entity_type"] == "ingredients":
                    tags = _normalized_tags(raw.get("tags"))
                    self._ensure_catalog_tags(tags, "ingredient")
                    procurement_params = {
                        "purchase_mode": raw.get("purchase_mode") or "regular",
                        "purchase_increment": raw.get("purchase_increment"),
                        "room_shelf_life_days": raw.get("room_shelf_life_days"),
                        "fridge_shelf_life_days": raw.get("fridge_shelf_life_days"),
                        "freezer_shelf_life_days": raw.get("freezer_shelf_life_days"),
                        "shelf_life_source": raw.get("shelf_life_source"),
                        "shelf_life_reviewed_at": raw.get("shelf_life_reviewed_at"),
                    }
                    if existing:
                        ingredient_id = existing["id"]
                        current_unit = self.session.execute(
                            text("SELECT default_unit FROM ingredients WHERE id=:id"),
                            {"id": ingredient_id},
                        ).scalar_one()
                        if current_unit != raw["default_unit"]:
                            referenced = self.session.execute(
                                text(
                                    """SELECT 1 FROM dish_ingredients WHERE ingredient_id=:id
                                       UNION ALL SELECT 1 FROM meal_ingredients WHERE ingredient_id=:id
                                       UNION ALL SELECT 1 FROM shopping_lists WHERE ingredient_id=:id
                                       LIMIT 1"""
                                ),
                                {"id": ingredient_id},
                            ).first()
                            if referenced:
                                raise ValidationAppError(
                                    f"Không thể đổi default_unit của nguyên liệu {raw['name']} đang được tham chiếu"
                                )
                        self.session.execute(
                            text("""UPDATE ingredients SET code=:code, name=:name, food_group=CAST(:group AS food_group),
                                   default_unit=:unit, grams_per_unit=:grams, tags=CAST(:tags AS jsonb),
                                   is_active=:active WHERE id=:id"""),
                            {"id": ingredient_id, "code": next_code, "name": raw["name"], "group": raw["food_group"],
                             "unit": raw["default_unit"], "grams": raw["grams_per_unit"], "tags": _json(tags),
                             "active": raw["is_active"]},
                        )
                        if raw.get("procurement_provided"):
                            self.session.execute(
                                text(
                                    """UPDATE ingredients SET
                                       purchase_mode=CAST(:purchase_mode AS ingredient_purchase_mode),
                                       purchase_increment=:purchase_increment,
                                       room_shelf_life_days=:room_shelf_life_days,
                                       fridge_shelf_life_days=:fridge_shelf_life_days,
                                       freezer_shelf_life_days=:freezer_shelf_life_days,
                                       shelf_life_source=:shelf_life_source,
                                       shelf_life_reviewed_at=:shelf_life_reviewed_at
                                       WHERE id=:id"""
                                ),
                                {"id": ingredient_id, **procurement_params},
                            )
                        updated += 1
                    else:
                        inserted = self.session.execute(
                            text("""INSERT INTO ingredients
                                   (code, name, food_group, default_unit, grams_per_unit, tags, is_active,
                                    purchase_mode, purchase_increment, room_shelf_life_days,
                                    fridge_shelf_life_days, freezer_shelf_life_days,
                                    shelf_life_source, shelf_life_reviewed_at)
                                   VALUES (:code, :name, CAST(:group AS food_group), :unit, :grams,
                                            CAST(:tags AS jsonb), :active,
                                            CAST(:purchase_mode AS ingredient_purchase_mode), :purchase_increment,
                                            :room_shelf_life_days, :fridge_shelf_life_days,
                                            :freezer_shelf_life_days, :shelf_life_source,
                                            :shelf_life_reviewed_at) RETURNING id"""),
                            {"code": next_code, "name": raw["name"], "group": raw["food_group"], "unit": raw["default_unit"],
                             "grams": raw["grams_per_unit"], "tags": _json(tags), "active": raw["is_active"],
                             **procurement_params},
                        ).first()
                        ingredient_id = inserted.id
                        created += 1
                    if raw.get("nutrition") is not None:
                        self.session.execute(
                            text("""INSERT INTO nutrition_facts
                                   (ingredient_id, calories, protein_g, carbs_g, fat_g, fiber_g)
                                   VALUES (:ingredient_id, :calories, :protein_g, :carbs_g, :fat_g, :fiber_g)
                                   ON CONFLICT (ingredient_id) DO UPDATE SET calories=EXCLUDED.calories,
                                   protein_g=EXCLUDED.protein_g, carbs_g=EXCLUDED.carbs_g,
                                   fat_g=EXCLUDED.fat_g, fiber_g=EXCLUDED.fiber_g"""),
                            {"ingredient_id": ingredient_id, **raw["nutrition"]},
                        )
                    if raw.get("price") is not None:
                        self.session.execute(
                            text("""INSERT INTO price_snapshots
                                   (ingredient_id, price, unit, price_per_default_unit, source)
                                   VALUES (:ingredient_id, :price, :unit, :price_per_default_unit, :source)"""),
                            {"ingredient_id": ingredient_id, **raw["price"]},
                        )
                else:
                    tags = _normalized_tags(raw.get("tags"))
                    self._ensure_catalog_tags(tags, "dish")
                    ingredient_payload = []
                    for item in raw.get("ingredients", []):
                        ingredient_id = _as_optional_id(item.get("ingredient_id"))
                        found = self.session.execute(
                            text("SELECT id FROM ingredients WHERE id=:id"), {"id": ingredient_id}
                        ).first() if ingredient_id else None
                        if not found:
                            raise ValidationAppError(f"Không tìm thấy nguyên liệu trong món {raw['name']}")
                        ingredient_payload.append({
                            "ingredient_id": found.id,
                            "quantity": float(item.get("quantity") or 0),
                            "unit": str(item.get("unit") or "").strip(),
                            "max_extra_quantity": float(item.get("max_extra_quantity") or 0),
                            "extra_step_quantity": float(item["extra_step_quantity"])
                            if item.get("extra_step_quantity") is not None else None,
                        })
                    if existing:
                        dish_id = existing["id"]
                        self.session.execute(
                            text("""UPDATE dishes SET code=:code, name=:name, dish_type=CAST(:dtype AS dish_type),
                                   cooking_method=CAST(:method AS cooking_method), description=:description,
                                   instructions=:instructions, tags=CAST(:tags AS jsonb), is_active=:active
                                   WHERE id=:id"""),
                            {"id": dish_id, "code": next_code, "name": raw["name"], "dtype": raw["dish_type"],
                             "method": raw.get("cooking_method"), "description": raw.get("description"),
                             "instructions": raw.get("instructions"), "tags": _json(tags),
                             "active": raw.get("is_active", True)},
                        )
                        self.session.execute(text("DELETE FROM dish_ingredients WHERE dish_id=:id"), {"id": dish_id})
                        updated += 1
                    else:
                        inserted = self.session.execute(
                            text("""INSERT INTO dishes
                                   (code, name, dish_type, cooking_method, description, instructions, tags, is_active)
                                   VALUES (:code, :name, CAST(:dtype AS dish_type), CAST(:method AS cooking_method),
                                   :description, :instructions, CAST(:tags AS jsonb), :active) RETURNING id"""),
                            {"code": next_code, "name": raw["name"], "dtype": raw["dish_type"], "method": raw.get("cooking_method"),
                             "description": raw.get("description"), "instructions": raw.get("instructions"),
                             "tags": _json(tags), "active": raw.get("is_active", True)},
                        ).first()
                        dish_id = inserted.id
                        created += 1
                    for item in ingredient_payload:
                        if item["quantity"] <= 0:
                            raise ValidationAppError(f"Định lượng trong món {raw['name']} phải lớn hơn 0")
                        if not item["unit"]:
                            raise ValidationAppError(f"Thiếu đơn vị trong món {raw['name']}")
                        self.session.execute(
                            text("""INSERT INTO dish_ingredients
                                   (dish_id, ingredient_id, quantity, unit,
                                    max_extra_quantity, extra_step_quantity)
                                   VALUES (:dish_id, :ingredient_id, :quantity, :unit,
                                           :max_extra_quantity, :extra_step_quantity)"""),
                            {"dish_id": dish_id, **item},
                        )
            summary = {"created": created, "updated": updated, "skipped": skipped}
            self.session.execute(
                text("UPDATE import_jobs SET status='committed', completed_at=NOW() WHERE id=:id"),
                {"id": job_id},
            )
            self._audit(actor_id, "import", job_data["entity_type"], job_id, after=summary)
            self.session.commit()
            return {"job_id": job_id, "status": "committed", **summary}
        except Exception:
            self.session.rollback()
            raise

    def list_import_jobs(self, limit: int, offset: int) -> dict[str, Any]:
        total = self.session.execute(text("SELECT COUNT(*) FROM import_jobs")).scalar_one()
        rows = self.session.execute(
            text("""SELECT id, entity_type, filename, status, total_rows, valid_rows,
                          error_count, created_by, created_at, completed_at
                   FROM import_jobs ORDER BY created_at DESC LIMIT :limit OFFSET :offset"""),
            {"limit": limit, "offset": offset},
        ).fetchall()
        return {"items": [_row_dict(r) for r in rows], "total": total, "limit": limit, "offset": offset}
